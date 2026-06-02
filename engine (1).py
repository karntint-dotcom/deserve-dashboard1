"""
engine.py — Deserve Dashboard (Hybrid edition)
- Sheet 1,2,3,5,6,12 → Excel formula (แก้ DATA แล้ว Ctrl+Alt+F9 update ทันที)
- Sheet 4,7,8,9,10,11 → Python computed (ข้อมูลแน่น วิเคราะห์ได้ลึก)
- Sheet 6,7,8,9,11 → มีคอลัมน์จำนวนชิ้น (unit) ที่ขายออก
"""
import io, warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gc
warnings.filterwarnings('ignore')

# ─── STYLES ──────────────────────────────────────────────────────────────────
T  = Side(style='thin',   color='D0D0D0')
M  = Side(style='medium', color='595959')
BT = Border(left=T, right=T, top=T, bottom=T)
BM = Border(left=M, right=M, top=M, bottom=M)

def F(h): return PatternFill("solid", fgColor=h)
def A(h="center", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def W(ws, r, c, val, bg=None, fg="000000", b=False, sz=9, fmt=None,
      h="center", wrap=False, border=BT, italic=False, merge_to=None):
    cc = ws.cell(row=r, column=c)
    cc.value = val
    cc.font = Font(name="Arial", bold=b, size=sz, color=fg, italic=italic)
    cc.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    cc.border = border
    if bg:  cc.fill = F(bg)
    if fmt: cc.number_format = fmt
    if merge_to:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    return cc

MTH = {1:'ม.ค.',2:'ก.พ.',3:'มี.ค.',4:'เม.ย.',5:'พ.ค.',
       6:'มิ.ย.',7:'ก.ค.',8:'ส.ค.',9:'ก.ย.',10:'ต.ค.',11:'พ.ย.',12:'ธ.ค.'}

STORE_CAT = {
    '🌟 STAR':       {'bg':'FFF0A0','fg':'7B5E00'},
    '📈 GROWTH':     {'bg':'C6EFCE','fg':'375623'},
    '🔄 RECOVERING': {'bg':'DEEAF1','fg':'1F4E79'},
    '📉 DECLINING':  {'bg':'FCE4D6','fg':'843C0C'},
    '⚠️ WARNING':    {'bg':'FFD7D7','fg':'C00000'},
    '😴 STABLE':     {'bg':'F2F2F2','fg':'595959'},
    '🔵 SPORADIC':   {'bg':'EBF3FB','fg':'1F4E79'},
    '⚡ INACTIVE':   {'bg':'F5F5F5','fg':'AAAAAA'},
}
PAIR_STYLE = {
    'GROWING':  ('C6EFCE','375623','📈 กำลังโต'),
    'STABLE':   ('F2F2F2','595959','➡️ คงที่'),
    'DECLINING':('FCE4D6','843C0C','📉 ลดลง'),
    'DROPPED':  ('FFD7D7','C00000','⛔ หยุดซื้อ'),
    'STOPPED':  ('FFE0B2','7B3F00','🔴 หายไป'),
    'SPORADIC': ('EBF3FB','1F4E79','💤 ไม่สม่ำเสมอ'),
    'ONE_TIME': ('FAFAFA','AAAAAA','1️⃣ ครั้งเดียว'),
}
CAT_CLR = {'Dog Food':'D6E4F7','Cat Food':'E8D5F5','Supplement':'D9F0E0',
           'Healthy Snack':'FFF2CC','Deserve Life':'FFE6D9',
           'RAW MATERIAL':'F2F2F2','Main SKU':'F2F2F2'}

# ─── DATA COLUMN REFS (source sheet layout) ──────────────────────────────────
CE='E'; CF='F'; CQ='Q'; CAK='AK'; CAR='AR'; CAS='AS'; CAT_='AT'; CAW='AW'; CAY='AY'
CBA='BA'; CBB='BB'; CBC='BC'; CBD='BD'
DAT=3   # first data row in DATA sheet

def _c(s):
    r=0
    for ch in s.upper(): r=r*26+(ord(ch)-64)
    return r

# ─── DATA LOADING ─────────────────────────────────────────────────────────────
def load_raw(file_bytes):
    buf = io.BytesIO(file_bytes)
    src = load_workbook(buf, data_only=False)
    buf.seek(0); c1 = pd.read_excel(buf, sheet_name='เงื่อนไข1', header=None)
    buf.seek(0); c2 = pd.read_excel(buf, sheet_name='เงื่อนไข2', header=None)
    buf.seek(0); df = pd.read_excel(buf, sheet_name='ข้อมูลเดือน 1-5', header=1)
    return src, df, c1, c2

def build_name_map(c1):
    n2s={}; cur=None
    for _,r_ in c1.iterrows():
        s=str(r_[0]).strip() if pd.notna(r_[0]) else ''
        c=str(r_[2]).strip() if pd.notna(r_[2]) else ''
        if s and s not in ('เงื่อนไข1','ชื่อร้าน','nan'): cur=s
        if c and c not in ('ชื่อลูกค้า','nan') and cur: n2s[c]=cur
    return n2s

def build_code_map(c2):
    cm={}
    for _,r_ in c2.iterrows():
        a=str(r_[2]).strip() if pd.notna(r_[2]) else ''
        b=str(r_[3]).strip() if pd.notna(r_[3]) else ''
        if b and b!='nan' and a and a!='nan': cm[b]=a
    return cm

def prep_df(df, n2s, code_map):
    df=df.copy()
    for col in ['ชื่อลูกค้า','รหัสลูกค้า','รหัสสินค้า','ชื่อสินค้า','หมวดหมู่','สถานะรายการ']:
        if col in df.columns: df[col]=df[col].astype(str).str.strip()
    df['ราคารวม']=pd.to_numeric(df.get('ราคารวม',pd.Series(dtype=float)),errors='coerce').fillna(0)
    df['จำนวน']  =pd.to_numeric(df.get('จำนวน',  pd.Series(dtype=float)),errors='coerce').fillna(0)
    df['date']   =pd.to_datetime(df.get('วันที่ทำรายการ'),dayfirst=True,errors='coerce')
    df['month']  =df['date'].dt.month
    df['ชื่อร้าน']=df['ชื่อลูกค้า'].map(n2s).fillna(df['ชื่อลูกค้า'])
    df['รหัสลูกค้า_norm']=df['รหัสลูกค้า'].map(lambda x:code_map.get(x,x))
    df['prod_grp']=df['รหัสสินค้า'].str[:6]
    df=df[df['สถานะรายการ']=='สำเร็จ'].copy()
    return df

def get_prod_maps(df):
    prod_names={}; prod_cat={}
    for grp,gdf in df.groupby('prod_grp'):
        nm=gdf['ชื่อสินค้า'].value_counts()
        for n in nm.index:
            if not any(ch.isdigit() for ch in str(n)[-8:]): prod_names[grp]=n; break
        if grp not in prod_names: prod_names[grp]=nm.index[0]
        prod_cat[grp]=gdf['หมวดหมู่'].mode()[0]
    return prod_names, prod_cat

def _linregress(x, y):
    xm,ym=x.mean(),y.mean(); ss_xy=((x-xm)*(y-ym)).sum(); ss_xx=((x-xm)**2).sum()
    sl=ss_xy/ss_xx if ss_xx!=0 else 0.0; ic=ym-sl*xm
    ss_tot=((y-ym)**2).sum()
    r2=max(0.0,1-((y-(sl*x+ic))**2).sum()/ss_tot) if ss_tot!=0 else 0.0
    return sl, ic, r2

def fc_linear(vals, months_act, months_fc):
    x=np.array(months_act,dtype=float); sl,ic,r2=_linregress(x,vals)
    return sl, r2, {m:max(0.0,sl*m+ic) for m in months_fc}

def wma3(vals): return float(max(0,np.average(vals[-3:],weights=[1,2,3])))

# ════════════════════════════════════════════════════════════════════════════
# HIDDEN SHEETS: COND1, COND2, STORE_LIST, CODE_LIST
# ════════════════════════════════════════════════════════════════════════════
def build_lookup_sheets(wb, src_wb, n2s, code_map, df):
    # Copy DATA sheet
    src_ws=src_wb['ข้อมูลเดือน 1-5']
    ws=wb.create_sheet("DATA"); ws.sheet_properties.tabColor="808080"
    max_row=src_ws.max_row
    for r in range(1,max_row+1):
        for c in range(1,src_ws.max_column+1):
            sc=src_ws.cell(r,c); dc=ws.cell(r,c); dc.value=sc.value
            if sc.has_style:
                dc.font=sc.font.copy(); dc.fill=sc.fill.copy()
                dc.border=sc.border.copy(); dc.alignment=sc.alignment.copy()
                dc.number_format=sc.number_format
    for cl,dim in src_ws.column_dimensions.items():
        ws.column_dimensions[cl].width=dim.width
    # Helper col headers
    for cl,lbl in [(CBA,"เดือน"),(CBB,"ชื่อร้าน"),(CBC,"รหัสลูกค้า_norm"),(CBD,"รหัสสินค้า_grp")]:
        W(ws,2,_c(cl),lbl,bg="1F4E79",fg="FFFFFF",b=True,sz=8)
        ws.column_dimensions[cl].width=14
    # Helper formulas
    for rn in range(DAT,max_row+1):
        ws.cell(rn,_c(CBA)).value=f'=IF({CAK}{rn}="สำเร็จ",IFERROR(MONTH(DATEVALUE({CQ}{rn})),MONTH({CQ}{rn})),"")'
        ws.cell(rn,_c(CBB)).value=f'=IFERROR(VLOOKUP({CE}{rn},COND1!A:B,2,0),{CE}{rn})'
        ws.cell(rn,_c(CBC)).value=f'=IFERROR(VLOOKUP({CF}{rn},COND2!A:B,2,0),{CF}{rn})'
        ws.cell(rn,_c(CBD)).value=f'=IF({CAR}{rn}<>"",LEFT({CAR}{rn},6),"")'
    ws.freeze_panes="A3"

    # COND1
    wc1=wb.create_sheet("COND1"); wc1.sheet_state='hidden'
    W(wc1,1,1,"ชื่อลูกค้า",bg="333333",fg="FFFFFF",b=True,sz=8)
    W(wc1,1,2,"ชื่อร้าน",  bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(cust,store) in enumerate(n2s.items(),2):
        wc1.cell(ri,1).value=cust; wc1.cell(ri,2).value=store
    wc1.column_dimensions['A'].width=40; wc1.column_dimensions['B'].width=30

    # COND2
    wc2=wb.create_sheet("COND2"); wc2.sheet_state='hidden'
    W(wc2,1,1,"รหัสลูกค้า_alt",  bg="333333",fg="FFFFFF",b=True,sz=8)
    W(wc2,1,2,"รหัสลูกค้า_norm", bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(alt,pri) in enumerate(code_map.items(),2):
        wc2.cell(ri,1).value=alt; wc2.cell(ri,2).value=pri

    # STORE_LIST & CODE_LIST (for ranking reference)
    stores=df.groupby('ชื่อร้าน')['ราคารวม'].sum().sort_values(ascending=False)
    codes=(df.groupby('รหัสลูกค้า_norm')
             .agg(ยอด=('ราคารวม','sum'),ร้าน=('ชื่อร้าน',lambda x:x.mode()[0]))
             .sort_values('ยอด',ascending=False))
    wsl=wb.create_sheet("STORE_LIST"); wsl.sheet_state='hidden'
    W(wsl,1,1,"ชื่อร้าน",bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,s in enumerate(stores.index,2): wsl.cell(ri,1).value=s
    wcl=wb.create_sheet("CODE_LIST"); wcl.sheet_state='hidden'
    W(wcl,1,1,"รหัส",bg="333333",fg="FFFFFF",b=True,sz=8)
    W(wcl,1,2,"ชื่อร้าน",bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(code,row) in enumerate(codes.iterrows(),2):
        wcl.cell(ri,1).value=code; wcl.cell(ri,2).value=row['ร้าน']

    return max_row, list(stores.index), list(codes.index), codes

# ════════════════════════════════════════════════════════════════════════════
# FORMULA SHEETS (1, 2, 3, 5, 6, 12)
# ════════════════════════════════════════════════════════════════════════════
def build_sheet1(wb, max_row, months):
    ws=wb.create_sheet("1_ภาพรวมยอดขาย"); ws.sheet_properties.tabColor="1F4E79"
    DR=max_row
    W(ws,1,1,"Deserve Dashboard – ภาพรวมยอดขาย",bg="1F4E79",fg="FFFFFF",b=True,sz=14,h="left",merge_to=7)
    W(ws,2,1,"⚡ Sheet 1,2,3,5,6,12 ใช้สูตร Excel — แก้ข้อมูลใน DATA แล้วกด Ctrl+Alt+F9",
      bg="FFF2CC",fg="7B5E00",sz=8,h="left",merge_to=7)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=14

    kpis=[
        ("ยอดขายรวม (บาท)",f'=SUMIF(DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ",DATA!{CAW}{DAT}:{CAW}{DR})','#,##0.00'),
        ("จำนวนรายการ",f'=COUNTIF(DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ")','#,##0'),
        ("จำนวนชื่อร้าน",f'=SUMPRODUCT(1/COUNTIF(DATA!{CBB}{DAT}:{CBB}{DR},DATA!{CBB}{DAT}:{CBB}{DR})*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ"))','#,##0'),
        ("จำนวนรหัสลูกค้า",f'=SUMPRODUCT(1/COUNTIF(DATA!{CBC}{DAT}:{CBC}{DR},DATA!{CBC}{DAT}:{CBC}{DR})*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ"))','#,##0'),
    ]
    for i,(lbl,f,fmt) in enumerate(kpis):
        W(ws,4,i*2+1,lbl,bg="2E75B6",fg="FFFFFF",b=True,sz=9,merge_to=i*2+2)
        c=ws.cell(5,i*2+1); c.value=f
        c.font=Font(name="Arial",bold=True,size=13,color="1F4E79")
        c.alignment=A(); c.border=BT; c.fill=F("DEEAF1"); c.number_format=fmt
        ws.merge_cells(start_row=5,start_column=i*2+1,end_row=5,end_column=i*2+2)
    ws.row_dimensions[5].height=26

    for ci,h in enumerate(["เดือน","ยอดขาย (บาท)","% ของยอดรวม","จำนวนร้าน","เพิ่ม/ลด vs เดือนก่อน"],1):
        W(ws,7,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=9)
    ws.row_dimensions[7].height=18

    for ri_off,m in enumerate(months,1):
        ri=ri_off+7; alt="F2F2F2" if ri_off%2==0 else None
        W(ws,ri,1,MTH[m],bg=alt,b=True)
        c2=ws.cell(ri,2)
        c2.value=(f'=SUMPRODUCT((DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                  f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
        c2.font=Font(name="Arial",size=9); c2.alignment=A(); c2.border=BT; c2.number_format='#,##0.00'
        if alt: c2.fill=F(alt)
        c3=ws.cell(ri,3); c3.value=f'=B{ri}/B5'
        c3.font=Font(name="Arial",size=9); c3.alignment=A(); c3.border=BT; c3.number_format='0.0%'
        if alt: c3.fill=F(alt)
        c4=ws.cell(ri,4)
        c4.value=(f'=IFERROR(SUMPRODUCT(1/COUNTIFS(DATA!{CBB}{DAT}:{CBB}{DR},DATA!{CBB}{DAT}:{CBB}{DR},'
                  f'DATA!{CBA}{DAT}:{CBA}{DR},{m},DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ")'
                  f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")),0)')
        c4.font=Font(name="Arial",size=9); c4.alignment=A(); c4.border=BT; c4.number_format='#,##0'
        if alt: c4.fill=F(alt)
        if ri_off>1:
            c6=ws.cell(ri,5); c6.value=f'=(B{ri}-B{ri-1})/B{ri-1}'
            c6.font=Font(name="Arial",size=9); c6.alignment=A(); c6.border=BT
            c6.number_format='+0.0%;-0.0%;0.0%'
            if alt: c6.fill=F(alt)

    tr=8+len(months)
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True)
    c=ws.cell(tr,2); c.value="=B5"; c.font=Font(name="Arial",bold=True,color="FFFFFF")
    c.alignment=A(); c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0.00'
    W(ws,tr,3,"=1",bg="BDD7EE",b=True,fmt='0.0%')
    for i,w_ in enumerate([10,18,14,14,20],1): ws.column_dimensions[gc(i)].width=w_


def build_sheet2(wb, max_row, months, store_list):
    ws=wb.create_sheet("2_Ranking_ร้านค้า"); ws.sheet_properties.tabColor="2E75B6"; DR=max_row
    W(ws,1,1,"Ranking ยอดขายรายเดือน – ชื่อร้าน",bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=4+len(months))
    hdrs=['อันดับ','ชื่อร้าน']+[MTH[m] for m in months]+['รวม (บาท)','เฉลี่ย/เดือน','%ยอดรวม']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height=28
    for ri_off,store in enumerate(store_list,1):
        ri=ri_off+3; alt="F0F4FF" if ri_off%2==0 else None
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10))
        W(ws,ri,2,store,bg=alt,h="left",b=(ri_off<=5),sz=9)
        for ci,m in enumerate(months,3):
            c=ws.cell(ri,ci)
            c.value=(f'=SUMPRODUCT((DATA!{CBB}{DAT}:{CBB}{DR}=B{ri})'
                     f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                     f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            c.font=Font(name="Arial",size=8); c.alignment=A(); c.border=BT; c.number_format='#,##0'
            if alt: c.fill=F(alt)
        fm=gc(3); lm=gc(2+len(months))
        ct=ws.cell(ri,3+len(months)); ct.value=f'=SUM({fm}{ri}:{lm}{ri})'
        ct.font=Font(name="Arial",bold=(ri_off<=10),size=9); ct.alignment=A()
        ct.border=BT; ct.number_format='#,##0.00'; ct.fill=F("BDD7EE")
        ca=ws.cell(ri,4+len(months)); ca.value=f'=AVERAGE({fm}{ri}:{lm}{ri})'
        ca.font=Font(name="Arial",size=8); ca.alignment=A(); ca.border=BT; ca.number_format='#,##0'
        if alt: ca.fill=F(alt)
        cp=ws.cell(ri,5+len(months)); cp.value=f"={gc(3+len(months))}{ri}/'1_ภาพรวมยอดขาย'!B5"
        cp.font=Font(name="Arial",size=8); cp.alignment=A(); cp.border=BT; cp.number_format='0.00%'
        if alt: cp.fill=F(alt)
    tr=len(store_list)+4
    W(ws,tr,2,"รวมทั้งหมด",bg="2E75B6",fg="FFFFFF",b=True,h="left")
    for ci,m in enumerate(months,3):
        c=ws.cell(tr,ci); c.value=f'=SUM({gc(ci)}4:{gc(ci)}{tr-1})'
        c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.alignment=A()
        c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0'
    ct=ws.cell(tr,3+len(months)); ct.value="='1_ภาพรวมยอดขาย'!B5"
    ct.font=Font(name="Arial",bold=True,color="FFFFFF"); ct.alignment=A()
    ct.border=BM; ct.fill=F("1F4E79"); ct.number_format='#,##0.00'
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=30
    for i in range(3,3+len(months)+4): ws.column_dimensions[gc(i)].width=12
    ws.freeze_panes="C4"


def build_sheet3(wb, max_row, months, code_list, codes_df):
    ws=wb.create_sheet("3_Ranking_รหัสลูกค้า"); ws.sheet_properties.tabColor="375623"; DR=max_row
    W(ws,1,1,"Ranking ยอดขายรายเดือน – รหัสลูกค้า",bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=5+len(months))
    hdrs=['อันดับ','ชื่อร้าน','รหัสลูกค้า']+[MTH[m] for m in months]+['รวม','เฉลี่ย','%ยอดรวม']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height=28
    for ri_off,(code,row) in enumerate(codes_df.iterrows(),1):
        ri=ri_off+3; alt="F0F4FF" if ri_off%2==0 else None
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10))
        W(ws,ri,2,row['ร้าน'],bg=alt,h="left",sz=8)
        W(ws,ri,3,code,bg=alt,h="left",sz=8)
        for ci,m in enumerate(months,4):
            c=ws.cell(ri,ci)
            c.value=(f'=SUMPRODUCT((DATA!{CBC}{DAT}:{CBC}{DR}=C{ri})'
                     f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                     f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            c.font=Font(name="Arial",size=8); c.alignment=A(); c.border=BT; c.number_format='#,##0'
            if alt: c.fill=F(alt)
        fm=gc(4); lm=gc(3+len(months))
        ct=ws.cell(ri,4+len(months)); ct.value=f'=SUM({fm}{ri}:{lm}{ri})'
        ct.font=Font(name="Arial",bold=(ri_off<=10),size=9); ct.alignment=A()
        ct.border=BT; ct.number_format='#,##0.00'; ct.fill=F("BDD7EE")
        ca=ws.cell(ri,5+len(months)); ca.value=f'=AVERAGE({fm}{ri}:{lm}{ri})'
        ca.font=Font(name="Arial",size=8); ca.alignment=A(); ca.border=BT; ca.number_format='#,##0'
        if alt: ca.fill=F(alt)
        cp=ws.cell(ri,6+len(months)); cp.value=f"={gc(4+len(months))}{ri}/'1_ภาพรวมยอดขาย'!B5"
        cp.font=Font(name="Arial",size=8); cp.alignment=A(); cp.border=BT; cp.number_format='0.00%'
        if alt: cp.fill=F(alt)
    tr=len(code_list)+4
    W(ws,tr,3,"รวมทั้งหมด",bg="2E75B6",fg="FFFFFF",b=True,h="left")
    for ci,m in enumerate(months,4):
        c=ws.cell(tr,ci); c.value=f'=SUM({gc(ci)}4:{gc(ci)}{tr-1})'
        c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.alignment=A()
        c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0'
    ct=ws.cell(tr,4+len(months)); ct.value="='1_ภาพรวมยอดขาย'!B5"
    ct.font=Font(name="Arial",bold=True,color="FFFFFF"); ct.alignment=A()
    ct.border=BM; ct.fill=F("1F4E79"); ct.number_format='#,##0.00'
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=26; ws.column_dimensions['C'].width=32
    for i in range(4,4+len(months)+4): ws.column_dimensions[gc(i)].width=12
    ws.freeze_panes="D4"


def build_sheet5(wb, max_row, months, plc_target, plc_deadline, recco_target, recco_deadline):
    ws=wb.create_sheet("5_เป้าหมาย_PLC_Recco"); ws.sheet_properties.tabColor="C00000"; DR=max_row
    W(ws,1,1,"ติดตามเป้าหมายยอดขาย (สูตรดึงจาก DATA)",bg="C00000",fg="FFFFFF",b=True,sz=13,h="left",merge_to=12)

    def block(sc, store_label, target, deadline, use_bb=True):
        W(ws,3,sc,store_label,bg="2E75B6",fg="FFFFFF",b=True,sz=11,merge_to=sc+5)
        W(ws,4,sc,"เป้าหมาย (บาท)",bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
        tc=ws.cell(4,sc+2); tc.value=target
        tc.font=Font(name="Arial",bold=True,size=11); tc.alignment=A(); tc.border=BT
        tc.fill=F("FFFFFF"); tc.number_format='#,##0'
        ws.merge_cells(start_row=4,start_column=sc+2,end_row=4,end_column=sc+5)
        tgt=f"{gc(sc+2)}4"
        col_ref=CBB if use_bb else CE
        if use_bb:
            af=(f'=SUMPRODUCT((DATA!{CBB}{DAT}:{CBB}{DR}="{store_label}")'
                f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
        else:
            af=(f'=SUMPRODUCT((DATA!{CE}{DAT}:{CE}{DR}="{store_label}")'
                f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
        for ri2,(lbl2,val2,fmt2) in enumerate([
            ("ยอดปัจจุบัน",af,'#,##0.00'),
            ("ยอดที่เหลือ",f'=MAX(0,{tgt}-{gc(sc+2)}5)','#,##0.00'),
            ("% บรรลุเป้า",f'=IFERROR({gc(sc+2)}5/{tgt},0)','0.0%'),
            ("กำหนดเวลา",deadline,None),
        ],5):
            W(ws,ri2,sc,lbl2,bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
            c2=ws.cell(ri2,sc+2); c2.value=val2
            c2.font=Font(name="Arial",bold=True,size=11 if ri2!=8 else 9)
            c2.alignment=A(); c2.border=BT
            if fmt2: c2.number_format=fmt2
            vbg="FFF2CC" if ri2==7 else ("FCE4D6" if ri2==6 else "DEEAF1")
            c2.fill=F(vbg)
            ws.merge_cells(start_row=ri2,start_column=sc+2,end_row=ri2,end_column=sc+5)
        for ci2,h2 in enumerate(["เดือน","ยอดขาย","% ของเป้า"],sc):
            W(ws,10,ci2,h2,bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        for ri3,m in enumerate(months,11):
            if use_bb:
                mvf=(f'=SUMPRODUCT((DATA!{CBB}{DAT}:{CBB}{DR}="{store_label}")'
                     f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                     f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            else:
                mvf=(f'=SUMPRODUCT((DATA!{CE}{DAT}:{CE}{DR}="{store_label}")'
                     f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                     f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            alt2="F0F4FF" if ri3%2==0 else None
            W(ws,ri3,sc,MTH[m],bg=alt2,sz=8)
            mv=ws.cell(ri3,sc+1); mv.value=mvf; mv.font=Font(name="Arial",size=8)
            mv.alignment=A(); mv.border=BT; mv.number_format='#,##0.00'
            if alt2: mv.fill=F(alt2)
            pv=ws.cell(ri3,sc+2); pv.value=f'=IFERROR({gc(sc+1)}{ri3}/{tgt},0)'
            pv.font=Font(name="Arial",size=8); pv.alignment=A(); pv.border=BT; pv.number_format='0.0%'
            if alt2: pv.fill=F(alt2)
        sr=11+len(months)
        W(ws,sr,sc,"สะสม",bg="BDD7EE",b=True,sz=9)
        sc2=ws.cell(sr,sc+1); sc2.value=f'={gc(sc+2)}5'
        sc2.font=Font(name="Arial",bold=True); sc2.alignment=A(); sc2.border=BM
        sc2.fill=F("BDD7EE"); sc2.number_format='#,##0.00'
        sp2=ws.cell(sr,sc+2); sp2.value=f'=IFERROR({gc(sc+2)}5/{tgt},0)'
        sp2.font=Font(name="Arial",bold=True); sp2.alignment=A(); sp2.border=BM
        sp2.fill=F("BDD7EE"); sp2.number_format='0.0%'

    block(1,"Pet Lover Centre",plc_target,plc_deadline,use_bb=True)
    block(8,"บริษัท เรคโค เพ็ท จำกัด",recco_target,recco_deadline,use_bb=False)
    for i in range(1,15): ws.column_dimensions[gc(i)].width=14


def build_sheet6(wb, max_row, df):
    ws=wb.create_sheet("6_ภาพรวมสินค้า"); ws.sheet_properties.tabColor="7030A0"; DR=max_row
    W(ws,1,1,"ภาพรวมยอดขายสินค้า – แยกหมวดหมู่",bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=6)
    cats=df.groupby('หมวดหมู่')['ราคารวม'].sum().sort_values(ascending=False).index.tolist()
    for ci,h in enumerate(["หมวดหมู่","ยอดขาย (บาท)","% ของยอดรวม","จำนวนชิ้น","จำนวน SKU"],1):
        W(ws,3,ci,h,bg="3B1F78",fg="FFFFFF",b=True,sz=9)
    ws.row_dimensions[3].height=22
    for ri_off,cat in enumerate(cats,1):
        ri=ri_off+3; cbg=CAT_CLR.get(cat,'F2F2F2')
        W(ws,ri,1,cat,bg=cbg,h="left",b=True,sz=9)
        # ยอดขาย (formula)
        cv=ws.cell(ri,2)
        cv.value=(f'=SUMPRODUCT((DATA!{CAY}{DAT}:{CAY}{DR}=A{ri})'
                  f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
        cv.font=Font(name="Arial",size=9); cv.alignment=A(); cv.border=BT; cv.number_format='#,##0.00'
        # %
        cp=ws.cell(ri,3); cp.value=f"=B{ri}/'1_ภาพรวมยอดขาย'!B5"
        cp.font=Font(name="Arial",size=9); cp.alignment=A(); cp.border=BT; cp.number_format='0.00%'
        # จำนวนชิ้น (formula) ← ใหม่
        cq=ws.cell(ri,4)
        cq.value=(f'=SUMPRODUCT((DATA!{CAY}{DAT}:{CAY}{DR}=A{ri})'
                  f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAT_}{DAT}:{CAT_}{DR})')
        cq.font=Font(name="Arial",size=9); cq.alignment=A(); cq.border=BT; cq.number_format='#,##0'
        # จำนวน SKU (Python computed — static)
        sku_count=df[df['หมวดหมู่']==cat]['prod_grp'].nunique()
        W(ws,ri,5,sku_count,sz=9)
    tr=len(cats)+4
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True,h="left")
    c=ws.cell(tr,2); c.value="='1_ภาพรวมยอดขาย'!B5"
    c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.alignment=A()
    c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0.00'
    W(ws,tr,3,"=1",bg="BDD7EE",b=True,fmt='0.0%')
    # จำนวนชิ้นรวม
    cq2=ws.cell(tr,4); cq2.value=f'=SUM(D4:D{tr-1})'
    cq2.font=Font(name="Arial",bold=True,color="FFFFFF"); cq2.alignment=A()
    cq2.border=BM; cq2.fill=F("2E75B6"); cq2.number_format='#,##0'
    for i,w_ in enumerate([18,16,14,13,11],1): ws.column_dimensions[gc(i)].width=w_


def build_sheet12(wb, max_row, months, monthly_targets):
    ws=wb.create_sheet("12_เป้าหมายรายเดือน"); ws.sheet_properties.tabColor="FF0000"; DR=max_row
    all_months=sorted(set(list(months)+list(monthly_targets.keys())))
    W(ws,1,1,"🎯 เป้าหมายยอดขายรายเดือน vs ยอดจริง (สูตรดึงจาก DATA)",
      bg="C00000",fg="FFFFFF",b=True,sz=13,h="left",merge_to=8)
    ws.row_dimensions[1].height=26
    W(ws,3,1,"📝 ตั้งค่าเป้าหมายรายเดือน (แก้ไขได้โดยตรง)",bg="2E75B6",fg="FFFFFF",b=True,sz=9,h="left",merge_to=8)
    for ci,m in enumerate(all_months,1):
        W(ws,4,ci,MTH[m],bg="1F4E79",fg="FFFFFF",b=True,sz=8)
        tc=ws.cell(5,ci); tc.value=monthly_targets.get(m,0)
        tc.font=Font(name="Arial",bold=True,size=10); tc.alignment=A()
        tc.border=BM; tc.fill=F("FFFDE7"); tc.number_format='#,##0'
    ws.row_dimensions[5].height=20
    W(ws,5,len(all_months)+1,"← แก้ตรงนี้ได้เลย",bg=None,fg="C00000",b=True,sz=8,h="left",border=Border())
    for ci,h in enumerate(["เดือน","เป้าหมาย","ยอดจริง","% Achievement","Gap","สถานะ","จำนวนร้าน"],1):
        W(ws,7,ci,h,bg="C00000",fg="FFFFFF",b=True,sz=9,wrap=True)
    ws.row_dimensions[7].height=28
    for ri_off,m in enumerate(all_months,1):
        ri=ri_off+7; alt="F9F9F9" if ri_off%2==0 else None
        W(ws,ri,1,MTH[m],bg=alt,b=True,sz=10)
        W(ws,ri,2,f'={gc(ri_off)}5',bg=alt,fmt='#,##0',sz=9)
        if m in months:
            c3=ws.cell(ri,3)
            c3.value=(f'=SUMPRODUCT((DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                      f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            c3.font=Font(name="Arial",size=9); c3.alignment=A(); c3.border=BT
            c3.number_format='#,##0.00'
            if alt: c3.fill=F(alt)
            c4=ws.cell(ri,4); c4.value=f'=IFERROR(C{ri}/B{ri},0)'
            c4.font=Font(name="Arial",bold=True,size=9); c4.alignment=A(); c4.border=BT; c4.number_format='0.0%'
            c5=ws.cell(ri,5); c5.value=f'=C{ri}-B{ri}'
            c5.font=Font(name="Arial",size=9); c5.alignment=A(); c5.border=BT; c5.number_format='+#,##0;-#,##0;0'
            c6=ws.cell(ri,6)
            c6.value=f'=IF(D{ri}>=1,"✅ ถึงเป้า",IF(D{ri}>=0.9,"🟡 ใกล้เป้า",IF(D{ri}>=0.7,"🟠 ต่ำกว่าเป้า","🔴 ต่ำมาก")))'
            c6.font=Font(name="Arial",bold=True,size=9); c6.alignment=A(); c6.border=BT
            c7=ws.cell(ri,7)
            c7.value=(f'=IFERROR(SUMPRODUCT(1/COUNTIFS(DATA!{CBB}{DAT}:{CBB}{DR},DATA!{CBB}{DAT}:{CBB}{DR},'
                      f'DATA!{CBA}{DAT}:{CBA}{DR},{m},DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ")'
                      f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")),0)')
            c7.font=Font(name="Arial",size=9); c7.alignment=A(); c7.border=BT
        else:
            for ci in range(3,8): W(ws,ri,ci,"—",bg="F5F5F5",fg="AAAAAA",sz=8,italic=True)
    tr=len(all_months)+8
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True)
    W(ws,tr,2,f'=SUM(B8:B{tr-1})',bg="BDD7EE",b=True,fmt='#,##0')
    c=ws.cell(tr,3); c.value=f'=SUM(C8:C{tr-1})'
    c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.alignment=A()
    c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0.00'
    W(ws,tr,4,f'=IFERROR(C{tr}/B{tr},0)',bg="BDD7EE",b=True,fmt='0.0%')
    W(ws,tr,5,f'=C{tr}-B{tr}',bg="BDD7EE",b=True,fmt='+#,##0;-#,##0;0')
    for i in range(1,9): ws.column_dimensions[gc(i)].width=15
    ws.freeze_panes="A8"

# ════════════════════════════════════════════════════════════════════════════
# PYTHON SHEETS (4, 7, 8, 9, 10, 11) — hardcoded values
# ════════════════════════════════════════════════════════════════════════════
def make_pivot(df, group_col, months, val='ราคารวม'):
    pm=df.groupby([group_col,'month'])[val].sum().unstack(fill_value=0)
    for m in months:
        if m not in pm.columns: pm[m]=0
    return pm[months]


def build_sheet4(wb, df, grand_total, months):
    ws=wb.create_sheet("4_ยอดขายฟาร์ม"); ws.sheet_properties.tabColor="375623"
    farm=df[df['รหัสลูกค้า'].str.startswith('ฟาร์ม')].copy()
    ft=farm['ราคารวม'].sum(); nf=farm['รหัสลูกค้า'].nunique()
    W(ws,1,1,"ยอดขายลูกค้าฟาร์ม",bg="375623",fg="FFFFFF",b=True,sz=13,h="left",merge_to=5+len(months))
    W(ws,2,1,f"จำนวน {nf} ฟาร์ม | ยอดรวม {ft:,.2f} บาท | {ft/grand_total:.2%} ของยอดทั้งหมด",
      bg="E2EFDA",fg="375623",b=True,sz=10,h="left",merge_to=5+len(months))
    fm=make_pivot(farm,'รหัสลูกค้า',months)
    fm['รวม']=fm.sum(axis=1); fm['เฉลี่ย']=fm[months].mean(axis=1)
    fm['%']=fm['รวม']/grand_total
    fm['เจ้าของ']=fm.index.map(farm.groupby('รหัสลูกค้า')['ชื่อลูกค้า'].first())
    fm=fm.sort_values('รวม',ascending=False)
    for ci,h in enumerate(['อันดับ','ฟาร์ม (รหัสลูกค้า)','เจ้าของ']+[MTH[m] for m in months]+['รวม','เฉลี่ย','%'],1):
        W(ws,4,ci,h,bg="375623",fg="FFFFFF",b=True,sz=8)
    ws.row_dimensions[4].height=22
    for ri_off,(fc,row) in enumerate(fm.iterrows(),1):
        ri=ri_off+4; alt="F0FFF0" if ri_off%2==0 else None
        W(ws,ri,1,ri_off,bg=alt,b=True); W(ws,ri,2,fc,bg=alt,h="left",b=True,sz=9)
        W(ws,ri,3,row.get('เจ้าของ',''),bg=alt,h="left",sz=8)
        for ci,m in enumerate(months,4):
            v=row[m]; W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt,fmt='#,##0',sz=8)
        W(ws,ri,4+len(months),row['รวม'],bg="A9D18E",b=True,fmt='#,##0.00')
        W(ws,ri,5+len(months),row['เฉลี่ย'],bg=alt,fmt='#,##0',sz=8)
        W(ws,ri,6+len(months),row['%'],bg=alt,fmt='0.00%',sz=8)
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=34; ws.column_dimensions['C'].width=26
    for i in range(4,4+len(months)+4): ws.column_dimensions[gc(i)].width=12


def build_sheet7(wb, df, grand_total, months, prod_names, prod_cat):
    ws=wb.create_sheet("7_Ranking_สินค้า"); ws.sheet_properties.tabColor="7030A0"
    W(ws,1,1,"Ranking ยอดขายสินค้า (กลุ่มรหัส 6 ตัว)",bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=6+len(months)+3)
    pm=make_pivot(df,'prod_grp',months,'ราคารวม')
    qm=make_pivot(df,'prod_grp',months,'จำนวน')    # ← จำนวนชิ้น
    pm['รวมบาท']=pm.sum(axis=1); pm['รวมชิ้น']=qm.sum(axis=1)
    pm['เฉลี่ยบาท']=pm[months].mean(axis=1); pm['เฉลี่ยชิ้น']=qm[months].mean(axis=1)
    pm['%']=pm['รวมบาท']/grand_total
    pm['ชื่อ']=pm.index.map(prod_names); pm['หมวด']=pm.index.map(prod_cat)
    pm=pm.sort_values('รวมบาท',ascending=False)
    hdrs=['อันดับ','รหัส','ชื่อสินค้า','หมวดหมู่']+[MTH[m] for m in months]+['รวม (บาท)','รวม (ชิ้น)','เฉลี่ย/เดือน (บาท)','เฉลี่ย/เดือน (ชิ้น)','%ยอดรวม']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="3B1F78",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height=36
    for ri_off,(code,row) in enumerate(pm.iterrows(),1):
        ri=ri_off+3; alt="F5F0FF" if ri_off%2==0 else None; ws.row_dimensions[ri].height=14
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10)); W(ws,ri,2,code,bg=alt,sz=8)
        W(ws,ri,3,row['ชื่อ'],bg=alt,h="left",sz=8,wrap=True)
        W(ws,ri,4,row['หมวด'],bg=CAT_CLR.get(row['หมวด'],'F2F2F2'),sz=8)
        for ci,m in enumerate(months,5):
            v=row[m]; W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt,fmt='#,##0',sz=8)
        ci_next=5+len(months)
        W(ws,ri,ci_next,  row['รวมบาท'],  bg="E8D5F5",b=True,fmt='#,##0.00'); ci_next+=1
        W(ws,ri,ci_next,  row['รวมชิ้น'], bg="D9F0E0",b=True,fmt='#,##0');    ci_next+=1   # ← ใหม่
        W(ws,ri,ci_next,  row['เฉลี่ยบาท'],bg=alt,fmt='#,##0',sz=8);         ci_next+=1
        W(ws,ri,ci_next,  row['เฉลี่ยชิ้น'],bg=alt,fmt='#,##0',sz=8);        ci_next+=1   # ← ใหม่
        W(ws,ri,ci_next,  row['%'],         bg=alt,fmt='0.00%',sz=8)
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=9
    ws.column_dimensions['C'].width=34; ws.column_dimensions['D'].width=13
    for i in range(5,5+len(months)+6): ws.column_dimensions[gc(i)].width=11
    ws.freeze_panes="E4"


def build_sheets_89(wb, df, grand_total, months, prod_names, prod_cat):
    for sheet_num,group_col,sname,tab,hbg in [
        (8,'ชื่อร้าน',       '8_สินค้า×Top10ร้าน', 'ED7D31','ED7D31'),
        (9,'รหัสลูกค้า_norm','9_สินค้า×Top10รหัส', 'FF0000','C00000'),
    ]:
        ws=wb.create_sheet(sname); ws.sheet_properties.tabColor=tab
        W(ws,1,1,f"Ranking สินค้า – Top10 {group_col}",bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=5+len(months)+2)
        top10=(df.groupby(group_col)['ราคารวม'].sum().sort_values(ascending=False).head(10).index.tolist())
        ri=3
        for rank_i,grp_val in enumerate(top10,1):
            grp_total=df[df[group_col]==grp_val]['ราคารวม'].sum()
            W(ws,ri,1,f"#{rank_i} {grp_val}  (รวม {grp_total:,.0f} บาท)",
              bg=hbg,fg="FFFFFF",b=True,sz=10,h="left",merge_to=5+len(months)+2); ri+=1
            hdr_bg="FCE4D6" if sheet_num==8 else "FCE9E9"
            for ci,h in enumerate(['รหัส','ชื่อสินค้า','หมวดหมู่']+[MTH[m] for m in months]+['รวม (บาท)','รวม (ชิ้น)','% ของกลุ่ม'],1):
                W(ws,ri,ci,h,bg=hdr_bg,b=True,sz=8); ri+=1   # ← เพิ่ม รวม (ชิ้น)
            sub=df[df[group_col]==grp_val]
            pm=make_pivot(sub,'prod_grp',months,'ราคารวม')
            qm=make_pivot(sub,'prod_grp',months,'จำนวน')
            pm['รวมบาท']=pm.sum(axis=1); pm['รวมชิ้น']=qm.sum(axis=1)
            pm=pm.sort_values('รวมบาท',ascending=False)
            for ri_off2,(pcode,prow) in enumerate(pm.iterrows(),1):
                alt2="F9F5FF" if ri_off2%2==0 else None; ws.row_dimensions[ri].height=13
                W(ws,ri,1,pcode,bg=alt2,sz=8)
                W(ws,ri,2,prod_names.get(pcode,pcode),bg=alt2,h="left",sz=8,wrap=True)
                W(ws,ri,3,prod_cat.get(pcode,''),bg=alt2,sz=8)
                for ci,m in enumerate(months,4):
                    v=prow.get(m,0); W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt2,fmt='#,##0',sz=8)
                ci_next=4+len(months)
                W(ws,ri,ci_next,prow['รวมบาท'],bg="BDD7EE",b=True,fmt='#,##0');     ci_next+=1
                W(ws,ri,ci_next,prow['รวมชิ้น'],bg="D9F0E0",b=True,fmt='#,##0');    ci_next+=1  # ← ใหม่
                W(ws,ri,ci_next,prow['รวมบาท']/grp_total if grp_total else 0,bg=alt2,fmt='0.0%',sz=8)
                ri+=1
            ri+=1
        ws.column_dimensions['A'].width=10; ws.column_dimensions['B'].width=34; ws.column_dimensions['C'].width=13
        for i in range(4,4+len(months)+4): ws.column_dimensions[gc(i)].width=11
        ws.freeze_panes="D2"


def build_sheet10(wb, df, grand_total, months):
    ACT=months; FCI=list(range(max(months)+1,13)) or list(range(6,13))
    ws=wb.create_sheet("10_Forecast_ร้านค้า"); ws.sheet_properties.tabColor="1F4E79"
    W(ws,1,1,f"📊 Forecast & วิเคราะห์ร้านค้า  |  ยอดจริง {MTH[min(ACT)]}–{MTH[max(ACT)]} + คาดการณ์ {MTH[min(FCI)] if FCI else ''}–{MTH[max(FCI)] if FCI else ''}",
      bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=3+len(ACT)+1+len(FCI)*2+2+3)
    for ci,(cat,st) in enumerate(STORE_CAT.items(),1):
        W(ws,2,ci,cat,bg=st['bg'],fg=st['fg'],b=True,sz=8)
    ws.row_dimensions[2].height=16
    sm=make_pivot(df,'ชื่อร้าน',ACT,'ราคารวม')
    store_rows=[]
    for store,row in sm.iterrows():
        v=row.values.astype(float); sl,r2,fc=fc_linear(v,ACT,FCI)
        w3=wma3(v); fc_ma={m:w3 for m in FCI}; active=sum(1 for x in v if x>0)
        peak_m=ACT[int(np.argmax(v))]; peak_v=float(np.max(v)); cur=float(v[-1])
        r2a=np.mean(v[len(v)//2:]); r2b=max(np.mean(v[:len(v)//2]),1)
        if   sl>8000 and r2>0.5:                cat='🌟 STAR'
        elif sl>2000 and r2>0.3:                cat='📈 GROWTH'
        elif r2a>r2b*1.2 and active>=3:         cat='🔄 RECOVERING'
        elif peak_v>0 and cur<peak_v*.55 and active>=3: cat='📉 DECLINING'
        elif sl<-3000 and active>=3:            cat='⚠️ WARNING'
        elif active<=2:                         cat='⚡ INACTIVE'
        elif active==3:                         cat='🔵 SPORADIC'
        else:                                   cat='😴 STABLE'
        actions={'🌟 STAR':"ขยาย SKU / volume discount",'📈 GROWTH':"ติดตาม / เสนอ bundle",
                 '🔄 RECOVERING':"โปร + ติดตาม 2 เดือน",'📉 DECLINING':f"ยอดตกจาก peak {MTH[peak_m]} → โทรหา",
                 '⚠️ WARNING':"เร่งด่วน! ตรวจสอบสาเหตุ",'⚡ INACTIVE':f"ไม่มียอดใน M{max(ACT)} → follow up",
                 '🔵 SPORADIC':"ซื้อไม่สม่ำเสมอ → สร้าง habit",'😴 STABLE':"รักษาฐาน / เสนอสินค้าใหม่"}
        store_rows.append({'ชื่อร้าน':store,'cat':cat,'action':actions.get(cat,''),
            **{f'M{m}':v[i] for i,m in enumerate(ACT)},
            'total':v.sum(),'pct':v.sum()/grand_total,
            'active':active,'peak_m':peak_m,'peak_v':peak_v,'cur':cur,
            'vs_peak':(cur-peak_v)/peak_v if peak_v>0 else 0,'slope':sl,'r2':r2,'wma3':w3,
            'fc_lin':fc,'fc_ma':fc_ma,'fc_lin_sum':sum(fc.values()),'fc_ma_sum':sum(fc_ma.values())})
    CAT_ORD=['🌟 STAR','📈 GROWTH','🔄 RECOVERING','📉 DECLINING','⚠️ WARNING','🔵 SPORADIC','😴 STABLE','⚡ INACTIVE']
    stores=pd.DataFrame(store_rows); stores['_so']=stores['cat'].map({v:i for i,v in enumerate(CAT_ORD)})
    stores=stores.sort_values(['_so','total'],ascending=[True,False])
    W(ws,3,1,"📌",bg="2E75B6",fg="FFFFFF",b=True)
    ci2=2
    for cat in CAT_ORD:
        sub=stores[stores['cat']==cat]
        if len(sub)>0:
            st=STORE_CAT.get(cat,{'bg':'F2F2F2','fg':'000000'})
            W(ws,3,ci2,f"{cat}: {len(sub)} ร้าน",bg=st['bg'],fg=st['fg'],b=True,sz=8,h="left"); ci2+=1
    ws.row_dimensions[3].height=16
    ws.row_dimensions[4].height=36
    HDRS=[('#','1F4E79'),('ชื่อร้าน','1F4E79'),('สถานะ','1F4E79')]
    for m in ACT: HDRS.append((f"{MTH[m]}\nจริง","2E75B6"))
    HDRS+=[('รวมจริง','1F4E79'),('Peak\nเดือน','595959'),('Peak\nมูลค่า','595959'),('vs Peak','595959')]
    for m in FCI: HDRS.append((f"🔵{MTH[m]}\nLinear","375623"))
    HDRS.append(('FC รวม\nLinear','1E5C1E'))
    for m in FCI: HDRS.append((f"🟠{MTH[m]}\nWMA","C55A11"))
    HDRS+=[('FC รวม\nWMA','7B2D00'),('slope/\nเดือน','595959'),('%ยอดรวม','595959'),('📋 แนะนำ','1F4E79')]
    for ci,(h,bg) in enumerate(HDRS,1): W(ws,4,ci,h,bg=bg,fg="FFFFFF",b=True,sz=8,wrap=True)
    for ri_off,(_,row) in enumerate(stores.iterrows(),1):
        r=ri_off+4; ws.row_dimensions[r].height=15
        st=STORE_CAT.get(row['cat'],{'bg':None,'fg':'000000'}); alt="F7FBFF" if ri_off%2==0 else None; ci=1
        W(ws,r,ci,ri_off,bg=alt,b=(ri_off<=10)); ci+=1
        W(ws,r,ci,row['ชื่อร้าน'],bg=alt,h="left",b=(row['cat'] in ['🌟 STAR','📈 GROWTH']),sz=9); ci+=1
        W(ws,r,ci,row['cat'],bg=st['bg'],fg=st['fg'],b=True,sz=8); ci+=1
        for m in ACT:
            v=row[f'M{m}']; mbg="FFF0A0" if m==row['peak_m'] and v>0 else("FFF0F0" if v==0 else alt)
            W(ws,r,ci,v,bg=mbg,fmt='#,##0',sz=8); ci+=1
        W(ws,r,ci,row['total'],bg="BDD7EE",b=True,fmt='#,##0.00'); ci+=1
        W(ws,r,ci,MTH.get(row['peak_m'],''),bg=alt,sz=8); ci+=1
        W(ws,r,ci,row['peak_v'],bg=alt,fmt='#,##0',sz=8); ci+=1
        vp=row['vs_peak']; vbg="FFD7D7" if vp<-0.4 else("FFF2CC" if vp<-0.15 else("E2EFDA" if vp>=0 else alt))
        W(ws,r,ci,vp,bg=vbg,fmt='+0%;-0%;0%',b=(abs(vp)>0.3),sz=8); ci+=1
        for m in FCI: W(ws,r,ci,row['fc_lin'][m],bg="E2EFDA" if row['fc_lin'][m]>0 else "FFF0F0",fmt='#,##0',sz=8,italic=True); ci+=1
        W(ws,r,ci,row['fc_lin_sum'],bg="A9D18E",b=True,fmt='#,##0.00'); ci+=1
        for m in FCI: W(ws,r,ci,row['fc_ma'][m],bg="FFF2E3",fmt='#,##0',sz=8,italic=True); ci+=1
        W(ws,r,ci,row['fc_ma_sum'],bg="F4B183",b=True,fmt='#,##0.00'); ci+=1
        W(ws,r,ci,row['slope'],bg="E2EFDA" if row['slope']>0 else "FFD7D7",fmt='+#,##0;-#,##0;0',sz=8); ci+=1
        W(ws,r,ci,row['pct'],bg=alt,fmt='0.00%',sz=8); ci+=1
        W(ws,r,ci,row['action'],bg=alt,h="left",sz=8,wrap=True)
    ws.column_dimensions['A'].width=5; ws.column_dimensions['B'].width=30; ws.column_dimensions['C'].width=18
    for i in range(4,4+len(ACT)+4): ws.column_dimensions[gc(i)].width=11
    for i in range(4+len(ACT)+4,4+len(ACT)+4+len(FCI)*2+2): ws.column_dimensions[gc(i)].width=9
    lc=4+len(ACT)+4+len(FCI)*2+2
    ws.column_dimensions[gc(lc)].width=9; ws.column_dimensions[gc(lc+1)].width=8; ws.column_dimensions[gc(lc+2)].width=30
    ws.freeze_panes=f"{gc(4+len(ACT))}5"


def build_sheet11(wb, df, grand_total, months, prod_names, prod_cat):
    ws=wb.create_sheet("11_Forecast_สินค้า"); ws.sheet_properties.tabColor="7030A0"
    W(ws,1,1,"📦 วิเคราะห์สินค้า × ร้านค้า  |  กำลังโต / เคยดีแล้วลด / หยุดซื้อ",
      bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=19)
    for ci,(k,(bg,fg,lbl)) in enumerate(PAIR_STYLE.items(),1):
        W(ws,2,ci,lbl,bg=bg,fg=fg,sz=8,b=True)
    ws.row_dimensions[2].height=15
    W(ws,3,1,"💡 header row = ยอดรวมสินค้าทุกร้าน  |  แถวย่อย = แต่ละร้านที่ซื้อสินค้านี้",
      bg="F3EEFF",fg="3B1F78",sz=8,h="left",merge_to=19)
    ws.row_dimensions[3].height=14
    # ← เพิ่ม รวม (ชิ้น) ในหัว
    P_HDRS=[('#','3B1F78'),('รหัส','3B1F78'),('ชื่อสินค้า','3B1F78'),('หมวดหมู่','3B1F78'),
            ('ชื่อร้าน','3B1F78'),('สถานะ','3B1F78')]
    for m in months: P_HDRS.append((f"{MTH[m]}\nจริง","2E75B6"))
    P_HDRS+=[('รวม (บาท)','1F4E79'),('รวม (ชิ้น)','375623'),
             ('%สินค้า','595959'),('Peak','595959'),
             ('Early\nAvg','595959'),('Late\nAvg','595959'),('Trend\n%','595959'),('💬 Insight','3B1F78')]
    for ci,(h,bg) in enumerate(P_HDRS,1): W(ws,4,ci,h,bg=bg,fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[4].height=36

    cross=df.groupby(['prod_grp','ชื่อร้าน','month'])['ราคารวม'].sum().unstack(fill_value=0)
    cross_q=df.groupby(['prod_grp','ชื่อร้าน','month'])['จำนวน'].sum().unstack(fill_value=0)
    for m in months:
        if m not in cross.columns:   cross[m]=0
        if m not in cross_q.columns: cross_q[m]=0
    cross=cross[months]; cross_q=cross_q[months]
    cross['total']=cross.sum(axis=1); cross['total_q']=cross_q.sum(axis=1)
    cross['early_avg']=cross[months[:3]].mean(axis=1)
    cross['late_avg'] =cross[months[-2:]].mean(axis=1)
    cross['trend_pct']=(cross['late_avg']-cross['early_avg'])/cross['early_avg'].replace(0,np.nan)
    cross['active']=(cross[months]>0).sum(axis=1)
    cross['peak_m']=cross[months].idxmax(axis=1); cross['peak_v']=cross[months].max(axis=1)
    cross['last_v']=cross[months[-1]]
    def cpair(row):
        tp=row['trend_pct']; ea=row['early_avg']; la=row['late_avg']
        if row['last_v']==0 and ea>5000: return 'STOPPED'
        if tp<-0.8 and ea>5000:          return 'DROPPED'
        if la>3000 and tp>0.3:            return 'GROWING'
        if row['active']<=1:              return 'ONE_TIME'
        if row['active']==2:              return 'SPORADIC'
        if tp<-0.2:                       return 'DECLINING'
        if tp>0.1:                        return 'GROWING'
        return 'STABLE'
    cross['pair_status']=cross.apply(cpair,axis=1)
    cross=cross.reset_index(); cross.columns.name=None
    cross['prod_name']=cross['prod_grp'].map(prod_names)
    cross=cross[cross['total']>1000].copy()
    prod_tot=df.groupby(['prod_grp','month'])['ราคารวม'].sum().unstack(fill_value=0)
    prod_tot_q=df.groupby(['prod_grp','month'])['จำนวน'].sum().unstack(fill_value=0)
    for m in months:
        if m not in prod_tot.columns:   prod_tot[m]=0
        if m not in prod_tot_q.columns: prod_tot_q[m]=0
    prod_tot=prod_tot[months]; prod_tot_q=prod_tot_q[months]
    prod_tot['total']=prod_tot.sum(axis=1); prod_tot['total_q']=prod_tot_q.sum(axis=1)
    prod_tot['early']=prod_tot[months[:3]].mean(axis=1); prod_tot['late']=prod_tot[months[-2:]].mean(axis=1)
    prod_tot['trend']=(prod_tot['late']-prod_tot['early'])/prod_tot['early'].replace(0,np.nan)
    prod_tot=prod_tot.sort_values('total',ascending=False)
    SORT_PS={'GROWING':0,'STABLE':1,'DECLINING':2,'DROPPED':3,'STOPPED':4,'SPORADIC':5,'ONE_TIME':6}
    ri=5; rank=0
    for pcode in prod_tot.index:
        if pcode not in prod_names: continue
        pname=prod_names[pcode]; pcat=prod_cat.get(pcode,'')
        prows=cross[cross['prod_grp']==pcode].copy()
        if len(prows)==0: continue
        ptotal=prod_tot.loc[pcode,'total']; ptotal_q=prod_tot.loc[pcode,'total_q']
        pearly=prod_tot.loc[pcode,'early']; plate=prod_tot.loc[pcode,'late']
        ptrend=prod_tot.loc[pcode,'trend']
        rank+=1
        pbg='E2EFDA' if ptrend>0.2 else('FCE4D6' if ptrend<-0.3 else 'DEEAF1')
        ptlbl='📈 GROWTH' if ptrend>0.2 else('📉 DECLINING' if ptrend<-0.3 else '➡️ STABLE')
        ws.row_dimensions[ri].height=17
        W(ws,ri,1,rank,bg=pbg,b=True,sz=9); W(ws,ri,2,pcode,bg=pbg,b=True,sz=9)
        W(ws,ri,3,pname,bg=pbg,b=True,sz=9,h="left",wrap=True)
        W(ws,ri,4,pcat,bg=CAT_CLR.get(pcat,'F2F2F2'),sz=8,b=True)
        W(ws,ri,5,f"รวม {len(prows)} ร้าน",bg=pbg,sz=8,italic=True)
        W(ws,ri,6,ptlbl,bg=pbg,b=True,sz=8)
        for ci2,m in enumerate(months,7):
            v=prod_tot.loc[pcode,m] if m in prod_tot.columns else 0
            W(ws,ri,ci2,v,bg=pbg,b=True,fmt='#,##0',sz=9)
        ci_n=7+len(months)
        W(ws,ri,ci_n,ptotal,bg=pbg,b=True,fmt='#,##0.00',sz=9);       ci_n+=1
        W(ws,ri,ci_n,ptotal_q,bg="D9F0E0",b=True,fmt='#,##0',sz=9);   ci_n+=1  # ← ใหม่
        W(ws,ri,ci_n,ptotal/grand_total,bg=pbg,fmt='0.00%',sz=8);      ci_n+=1
        peak_m2=prod_tot.loc[pcode,months].idxmax()
        W(ws,ri,ci_n,MTH.get(peak_m2,''),bg=pbg,sz=8);                 ci_n+=1
        W(ws,ri,ci_n,pearly,bg=pbg,fmt='#,##0',sz=8);                  ci_n+=1
        W(ws,ri,ci_n,plate,bg=pbg,fmt='#,##0',sz=8);                   ci_n+=1
        tbg="E2EFDA" if ptrend>0 else "FCE4D6"
        W(ws,ri,ci_n,ptrend if not np.isnan(ptrend) else 0,bg=tbg,fmt='+0%;-0%;0%',b=True,sz=8); ci_n+=1
        ng=len(prows[prows['pair_status']=='GROWING']); ns=len(prows[prows['pair_status'].isin(['STOPPED','DROPPED'])])
        ins=f"{ng} ร้านกำลังโต"+(f"  |  {ns} ร้านหยุดซื้อ" if ns>0 else "")
        W(ws,ri,ci_n,ins,bg=pbg,h="left",sz=8); ri+=1
        prows=prows.copy(); prows['_ps']=prows['pair_status'].map(SORT_PS).fillna(9)
        prows=prows.sort_values(['_ps','total'],ascending=[True,False])
        for _,pr in prows.iterrows():
            ws.row_dimensions[ri].height=14; ps=pr['pair_status']
            pb2,pf2,plbl2=PAIR_STYLE.get(ps,('F2F2F2','000000','')); alt2="FDFBFF" if ri%2==0 else None
            for ci2 in [1,2,3,4]: W(ws,ri,ci2,'',bg=alt2)
            W(ws,ri,5,pr['ชื่อร้าน'],bg=alt2,h="left",sz=8)
            W(ws,ri,6,plbl2,bg=pb2,fg=pf2,b=True,sz=8)
            for ci2,m in enumerate(months,7):
                v=pr.get(m,0); mbg="FFF0A0" if m==pr['peak_m'] and v>0 else("FFF0F0" if v==0 else alt2)
                W(ws,ri,ci2,v,bg=mbg,fmt='#,##0',sz=8)
            ci_n=7+len(months)
            W(ws,ri,ci_n,pr['total'],bg=alt2,b=(ps=='GROWING'),fmt='#,##0',sz=8);       ci_n+=1
            W(ws,ri,ci_n,pr['total_q'],bg="F0FFF0" if pr['total_q']>0 else "FFF0F0",fmt='#,##0',sz=8); ci_n+=1  # ← ใหม่
            pcp=pr['total']/ptotal if ptotal>0 else 0
            W(ws,ri,ci_n,pcp,bg=alt2,fmt='0.0%',sz=8); ci_n+=1
            W(ws,ri,ci_n,MTH.get(pr['peak_m'],''),bg=alt2,sz=8); ci_n+=1
            W(ws,ri,ci_n,pr['early_avg'],bg=alt2,fmt='#,##0',sz=8); ci_n+=1
            W(ws,ri,ci_n,pr['late_avg'],bg="FFF0F0" if pr['late_avg']==0 else alt2,fmt='#,##0',sz=8); ci_n+=1
            tp2=pr['trend_pct'] if not(isinstance(pr['trend_pct'],float) and np.isnan(pr['trend_pct'])) else 0
            W(ws,ri,ci_n,tp2,bg="E2EFDA" if tp2>0.1 else("FFD7D7" if tp2<-0.3 else alt2),fmt='+0%;-0%;0%',sz=8); ci_n+=1
            if ps=='GROWING':    ins2=f"↑ avg {pr['late_avg']:,.0f}฿ → ดัน SKU"
            elif ps=='STOPPED':  ins2=f"⛔ เคย avg {pr['early_avg']:,.0f}฿ → โทรหาด่วน"
            elif ps=='DROPPED':  ins2=f"⚠️ ลดจาก {pr['peak_v']:,.0f}฿ → ตรวจสอบ"
            elif ps=='DECLINING': ins2=f"↓ {tp2:.0%} → เสนอโปรฯ"
            elif ps=='SPORADIC':  ins2="ซื้อสลับ → สร้างความสม่ำเสมอ"
            elif ps=='ONE_TIME':  ins2="ซื้อครั้งเดียว → follow up"
            else:                 ins2="คงที่ → รักษาฐาน"
            W(ws,ri,ci_n,ins2,bg=alt2,h="left",sz=8,wrap=True); ri+=1
        ri+=1
    ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=8; ws.column_dimensions['C'].width=34
    ws.column_dimensions['D'].width=13; ws.column_dimensions['E'].width=28; ws.column_dimensions['F'].width=14
    for i in range(7,7+len(months)): ws.column_dimensions[gc(i)].width=10
    ci_base=7+len(months)
    ws.column_dimensions[gc(ci_base)].width=12; ws.column_dimensions[gc(ci_base+1)].width=11  # ← ชิ้น
    for i in range(ci_base+2,ci_base+7): ws.column_dimensions[gc(i)].width=10
    ws.column_dimensions[gc(ci_base+7)].width=32
    ws.freeze_panes="D5"

# ════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
def build_dashboard(file_bytes, plc_target=1_250_000, plc_deadline="30 มิถุนายน 2569",
                    recco_target=1_000_000, recco_deadline="31 ธันวาคม 2569",
                    monthly_targets=None, progress_cb=None):
    def upd(pct, msg):
        if progress_cb: progress_cb(pct, msg)

    upd(8,  "📥 โหลดข้อมูล...")
    src_wb, df_raw, c1, c2 = load_raw(file_bytes)
    n2s      = build_name_map(c1)
    code_map = build_code_map(c2)
    df       = prep_df(df_raw, n2s, code_map)
    months   = sorted(df['month'].dropna().unique().astype(int))
    grand_total = df['ราคารวม'].sum()
    prod_names, prod_cat = get_prod_maps(df)

    wb = Workbook(); wb.remove(wb.active)

    upd(15, "📋 สร้าง lookup tables + copy DATA...")
    max_row, store_list, code_list, codes_df = build_lookup_sheets(wb, src_wb, n2s, code_map, df)

    upd(25, "📊 ภาพรวมยอดขาย (formula)...")
    build_sheet1(wb, max_row, months)

    upd(32, "🏪 Ranking ร้านค้า (formula)...")
    build_sheet2(wb, max_row, months, store_list)

    upd(38, "🔑 Ranking รหัสลูกค้า (formula)...")
    build_sheet3(wb, max_row, months, code_list, codes_df)

    upd(43, "🌾 ยอดขายฟาร์ม...")
    build_sheet4(wb, df, grand_total, months)

    upd(49, "🎯 เป้าหมาย PLC & Recco (formula)...")
    build_sheet5(wb, max_row, months, plc_target, plc_deadline, recco_target, recco_deadline)

    upd(55, "📦 ภาพรวมสินค้า (formula + จำนวนชิ้น)...")
    build_sheet6(wb, max_row, df)

    upd(62, "📋 Ranking สินค้า (+ จำนวนชิ้น)...")
    build_sheet7(wb, df, grand_total, months, prod_names, prod_cat)

    upd(70, "🔗 สินค้า × Top10 (+ จำนวนชิ้น)...")
    build_sheets_89(wb, df, grand_total, months, prod_names, prod_cat)

    upd(80, "🔮 Forecast ร้านค้า...")
    build_sheet10(wb, df, grand_total, months)

    upd(90, "📦 Forecast สินค้า × ร้าน (+ จำนวนชิ้น)...")
    build_sheet11(wb, df, grand_total, months, prod_names, prod_cat)

    if monthly_targets:
        upd(96, "🎯 เป้าหมายรายเดือน (formula)...")
        build_sheet12(wb, max_row, months, monthly_targets)

    upd(99, "💾 บันทึกไฟล์...")
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()
