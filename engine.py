"""
engine.py — Deserve Dashboard (Excel-formula edition)
ข้อมูลดิบ copy มาไว้ใน Sheet_DATA แล้วทุก sheet ใช้ Excel formula ดึงจากนั้น
แก้ข้อมูลใน sheet DATA → กด F9 → ทุกชีต update ทันที
"""
import io, warnings
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gc
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

warnings.filterwarnings('ignore')

# ─── STYLE HELPERS ───────────────────────────────────────────────────────────
T  = Side(style='thin',   color='D0D0D0')
M  = Side(style='medium', color='595959')
BT = Border(left=T, right=T, top=T, bottom=T)
BM = Border(left=M, right=M, top=M, bottom=M)

def F(h): return PatternFill("solid", fgColor=h)
def A(h="center", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def W(ws, r, c, val, bg=None, fg="000000", b=False, sz=9, fmt=None,
      h="center", wrap=False, border=BT, italic=False, merge_to=None):
    cc = ws.cell(row=r, column=c)
    cc.value = val
    cc.font = Font(name="Arial", bold=b, size=sz, color=fg, italic=italic)
    cc.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    cc.border = border
    if bg:  cc.fill = F(bg)
    if fmt: cc.number_format = fmt
    if merge_to:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    return cc

MTH = {1:'ม.ค.',2:'ก.พ.',3:'มี.ค.',4:'เม.ย.',5:'พ.ค.',
       6:'มิ.ย.',7:'ก.ค.',8:'ส.ค.',9:'ก.ย.',10:'ต.ค.',11:'พ.ย.',12:'ธ.ค.'}

# ─── DATA LAYOUT CONSTANTS ───────────────────────────────────────────────────
# ในชีต DATA (copy จาก source) — header row 2, data starts row 3
D = "DATA"          # sheet name
HDR = 2             # header row
DAT = 3             # first data row
# Max data row จะถูก set หลัง copy ข้อมูล
# คอลัมน์หลัก (ตาม source layout)
CE  = "E"   # ชื่อลูกค้า
CF  = "F"   # รหัสลูกค้า
CQ  = "Q"   # วันที่ทำรายการ
CAK = "AK"  # สถานะรายการ
CAR = "AR"  # รหัสสินค้า
CAS = "AS"  # ชื่อสินค้า
CAT = "AT"  # จำนวน
CAW = "AW"  # ราคารวม
CAY = "AY"  # หมวดหมู่

# Helper columns เพิ่มใน DATA sheet (ต่อจากคอลัมน์สุดท้าย = col 51)
# col 53 = BA = เดือน  (=MONTH(Q3))
# col 54 = BB = ชื่อร้าน (=XLOOKUP จาก เงื่อนไข1)
# col 55 = BC = รหัสลูกค้า_norm (=IFERROR(XLOOKUP จาก เงื่อนไข2, F3))
# col 56 = BD = รหัสสินค้า_grp (=LEFT(AR3,6))
CBA = "BA"  # เดือน
CBB = "BB"  # ชื่อร้าน (normalized)
CBC = "BC"  # รหัสลูกค้า_norm
CBD = "BD"  # รหัสสินค้า_grp (6 chars)

# ─── LOAD RAW DATA ──────────────────────────────────────────────────────────
def load_raw(file_bytes):
    buf = io.BytesIO(file_bytes)
    src = load_workbook(buf, data_only=False)
    buf.seek(0)
    # Load for Python processing (conditions)
    c1  = pd.read_excel(buf, sheet_name='เงื่อนไข1', header=None)
    buf.seek(0)
    c2  = pd.read_excel(buf, sheet_name='เงื่อนไข2', header=None)
    buf.seek(0)
    df  = pd.read_excel(buf, sheet_name='ข้อมูลเดือน 1-5', header=1)
    return src, df, c1, c2

def build_name_map(c1):
    n2s = {}; cur = None
    for _, r_ in c1.iterrows():
        s = str(r_[0]).strip() if pd.notna(r_[0]) else ''
        c = str(r_[2]).strip() if pd.notna(r_[2]) else ''
        if s and s not in ('เงื่อนไข1','ชื่อร้าน','nan'): cur = s
        if c and c not in ('ชื่อลูกค้า','nan') and cur: n2s[c] = cur
    return n2s

def build_code_map(c2):
    code_map = {}
    for _, r_ in c2.iterrows():
        c1v = str(r_[2]).strip() if pd.notna(r_[2]) else ''
        c2v = str(r_[3]).strip() if pd.notna(r_[3]) else ''
        if c2v and c2v != 'nan' and c1v and c1v != 'nan': code_map[c2v] = c1v
    return code_map

# ─── SHEET: DATA (ข้อมูลดิบ + helper cols) ──────────────────────────────────
def build_data_sheet(wb, src_wb, n2s, code_map):
    """Copy raw data sheet + add helper formula columns"""
    # Copy source sheet
    src_ws = src_wb['ข้อมูลเดือน 1-5']
    ws = wb.create_sheet("DATA")
    ws.sheet_properties.tabColor = "808080"

    max_row = src_ws.max_row
    max_col = src_ws.max_column

    # Copy all cells
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dst_cell = ws.cell(row=r, column=c)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font      = src_cell.font.copy()
                dst_cell.fill      = src_cell.fill.copy()
                dst_cell.border    = src_cell.border.copy()
                dst_cell.alignment = src_cell.alignment.copy()
                dst_cell.number_format = src_cell.number_format

    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    # ── Helper column headers (row 2) ──
    for c_letter, label in [(CBA,"เดือน"),(CBB,"ชื่อร้าน"),(CBC,"รหัสลูกค้า_norm"),(CBD,"รหัสสินค้า_grp")]:
        W(ws, HDR, _col_num(c_letter), label, bg="1F4E79", fg="FFFFFF", b=True, sz=8)
        ws.column_dimensions[c_letter].width = 14

    # ── Helper formulas for each data row ──
    # Build lookup tables as Python dicts → write to hidden lookup sheets
    # then reference with IFERROR(VLOOKUP)

    # Build เงื่อนไข1 flat table: col A=ชื่อลูกค้า, col B=ชื่อร้าน
    # Build เงื่อนไข2 flat table: col A=รหัสลูกค้า_alt, col B=รหัสลูกค้า_primary

    for row_num in range(DAT, max_row + 1):
        # BA = เดือน
        ws.cell(row=row_num, column=_col_num(CBA)).value = \
            f'=IF({CAK}{row_num}="สำเร็จ",IFERROR(MONTH(DATEVALUE({CQ}{row_num})),MONTH({CQ}{row_num})),"")'
        ws.cell(row=row_num, column=_col_num(CBA)).number_format = '0'

        # BB = ชื่อร้าน → IFERROR(VLOOKUP(E, COND1!A:B, 2, 0), E)
        ws.cell(row=row_num, column=_col_num(CBB)).value = \
            f'=IFERROR(VLOOKUP({CE}{row_num},COND1!A:B,2,0),{CE}{row_num})'

        # BC = รหัสลูกค้า_norm → IFERROR(VLOOKUP(F, COND2!A:B, 2, 0), F)
        ws.cell(row=row_num, column=_col_num(CBC)).value = \
            f'=IFERROR(VLOOKUP({CF}{row_num},COND2!A:B,2,0),{CF}{row_num})'

        # BD = รหัสสินค้า_grp → LEFT(AR, 6)
        ws.cell(row=row_num, column=_col_num(CBD)).value = \
            f'=IF({CAR}{row_num}<>"",LEFT({CAR}{row_num},6),"")'

    ws.freeze_panes = "A3"
    return max_row

def _col_num(letter):
    """Convert column letter(s) to number"""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result

# ─── SHEET: COND1 (lookup table: ชื่อลูกค้า → ชื่อร้าน) ────────────────────
def build_cond1_sheet(wb, n2s):
    ws = wb.create_sheet("COND1")
    ws.sheet_properties.tabColor = "595959"
    ws.sheet_state = 'hidden'
    W(ws,1,1,"ชื่อลูกค้า",bg="333333",fg="FFFFFF",b=True,sz=8)
    W(ws,1,2,"ชื่อร้าน",  bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(cust,store) in enumerate(n2s.items(),2):
        ws.cell(row=ri,column=1).value = cust
        ws.cell(row=ri,column=2).value = store
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    return len(n2s) + 1   # last row

# ─── SHEET: COND2 (lookup table: รหัสลูกค้า_alt → primary) ─────────────────
def build_cond2_sheet(wb, code_map):
    ws = wb.create_sheet("COND2")
    ws.sheet_properties.tabColor = "595959"
    ws.sheet_state = 'hidden'
    W(ws,1,1,"รหัสลูกค้า_alt",  bg="333333",fg="FFFFFF",b=True,sz=8)
    W(ws,1,2,"รหัสลูกค้า_norm", bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(alt,primary) in enumerate(code_map.items(),2):
        ws.cell(row=ri,column=1).value = alt
        ws.cell(row=ri,column=2).value = primary
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35

# ─── SHEET: STORE_LIST (unique store names + codes for ranking) ──────────────
def build_store_list_sheet(wb, df, n2s, code_map):
    """Hidden sheet listing unique stores and codes — used by ranking sheets"""
    # Compute from Python (this is static reference data, not recalculated)
    df2 = df.copy()
    for col in ['ชื่อลูกค้า','รหัสลูกค้า']:
        df2[col] = df2[col].astype(str).str.strip()
    df2['ชื่อร้าน'] = df2['ชื่อลูกค้า'].map(n2s).fillna(df2['ชื่อลูกค้า'])
    df2['รหัสลูกค้า_norm'] = df2['รหัสลูกค้า'].map(lambda x: code_map.get(x,x))
    df2['ราคารวม'] = pd.to_numeric(df2['ราคารวม'], errors='coerce').fillna(0)
    df2 = df2[df2['สถานะรายการ']=='สำเร็จ']

    stores = df2.groupby('ชื่อร้าน')['ราคารวม'].sum().sort_values(ascending=False)
    codes  = (df2.groupby('รหัสลูกค้า_norm')
                 .agg(ยอด=('ราคารวม','sum'), ร้าน=('ชื่อร้าน', lambda x: x.mode()[0]))
                 .sort_values('ยอด', ascending=False))

    # Sheet: STORE_LIST
    ws = wb.create_sheet("STORE_LIST")
    ws.sheet_state = 'hidden'
    W(ws,1,1,"ชื่อร้าน",bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(store,_) in enumerate(stores.items(),2):
        ws.cell(row=ri,column=1).value = store
    ws.column_dimensions['A'].width = 30

    # Sheet: CODE_LIST
    wc = wb.create_sheet("CODE_LIST")
    wc.sheet_state = 'hidden'
    W(wc,1,1,"รหัสลูกค้า",bg="333333",fg="FFFFFF",b=True,sz=8)
    W(wc,1,2,"ชื่อร้าน",  bg="333333",fg="FFFFFF",b=True,sz=8)
    for ri,(code,row) in enumerate(codes.iterrows(),2):
        wc.cell(row=ri,column=1).value = code
        wc.cell(row=ri,column=2).value = row['ร้าน']
    wc.column_dimensions['A'].width = 32; wc.column_dimensions['B'].width = 28

    return list(stores.index), list(codes.index), codes

# ─── SHEET 1: ภาพรวมยอดขาย ──────────────────────────────────────────────────
def build_sheet1(wb, max_row, months):
    ws = wb.create_sheet("1_ภาพรวมยอดขาย")
    ws.sheet_properties.tabColor = "1F4E79"
    DR = max_row   # last data row in DATA sheet

    W(ws,1,1,"Deserve Dashboard – ภาพรวมยอดขาย",
      bg="1F4E79",fg="FFFFFF",b=True,sz=14,h="left",merge_to=7)
    W(ws,2,1,"⚡ ทุกตัวเลขในไฟล์นี้คำนวณจาก sheet DATA โดยตรง — แก้ข้อมูลใน DATA แล้วกด Ctrl+Alt+F9",
      bg="FFF2CC",fg="7B5E00",b=False,sz=8,h="left",merge_to=7)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=16

    # KPI row
    kpis = [
        ("ยอดขายรวม (บาท)",
         f'=SUMIF(DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ",DATA!{CAW}{DAT}:{CAW}{DR})'),
        ("จำนวนรายการ",
         f'=COUNTIF(DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ")'),
        ("จำนวนชื่อร้าน",
         f'=SUMPRODUCT(1/COUNTIF(DATA!{CBB}{DAT}:{CBB}{DR},DATA!{CBB}{DAT}:{CBB}{DR})*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ"))'),
        ("จำนวนรหัสลูกค้า",
         f'=SUMPRODUCT(1/COUNTIF(DATA!{CBC}{DAT}:{CBC}{DR},DATA!{CBC}{DAT}:{CBC}{DR})*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ"))'),
    ]
    fmts = ['#,##0.00','#,##0','#,##0','#,##0']
    for i,(lbl,formula) in enumerate(kpis):
        W(ws,4,i*2+1,lbl,bg="2E75B6",fg="FFFFFF",b=True,sz=9,merge_to=i*2+2)
        c=ws.cell(row=5,column=i*2+1)
        c.value=formula; c.font=Font(name="Arial",bold=True,size=13,color="1F4E79")
        c.alignment=A(); c.border=BT; c.fill=F("DEEAF1")
        c.number_format=fmts[i]
        ws.merge_cells(start_row=5,start_column=i*2+1,end_row=5,end_column=i*2+2)
    ws.row_dimensions[5].height=26

    # Monthly table
    hdrs=["เดือน","ยอดขาย (บาท)","% ของยอดรวม","จำนวนร้าน","จำนวนรหัส","เพิ่ม/ลด vs เดือนก่อน"]
    for ci,h in enumerate(hdrs,1): W(ws,7,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=9)
    ws.row_dimensions[7].height=18

    grand_ref = f"B{5}"   # KPI ยอดรวม cell

    for ri_off,m in enumerate(months,1):
        ri = ri_off + 7
        ws.row_dimensions[ri].height=16
        alt = "F2F2F2" if ri_off%2==0 else None

        # ยอดขาย = SUMPRODUCT month filter
        sales_f = (f'=SUMPRODUCT((DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                   f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")'
                   f'*DATA!{CAW}{DAT}:{CAW}{DR})')

        W(ws,ri,1,MTH[m],bg=alt,b=True)
        c=ws.cell(row=ri,column=2); c.value=sales_f
        c.font=Font(name="Arial",size=9); c.alignment=A(); c.border=BT
        c.number_format='#,##0.00'
        if alt: c.fill=F(alt)

        # % ของยอดรวม
        c3=ws.cell(row=ri,column=3)
        c3.value=f'=B{ri}/B5'; c3.font=Font(name="Arial",size=9)
        c3.alignment=A(); c3.border=BT; c3.number_format='0.0%'
        if alt: c3.fill=F(alt)

        # จำนวนร้าน
        c4=ws.cell(row=ri,column=4)
        c4.value=(f'=SUMPRODUCT(1/COUNTIFS(DATA!{CBB}{DAT}:{CBB}{DR},DATA!{CBB}{DAT}:{CBB}{DR},'
                  f'DATA!{CBA}{DAT}:{CBA}{DR},{m},'
                  f'DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ")'
                  f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                  f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ"))')
        c4.font=Font(name="Arial",size=9); c4.alignment=A(); c4.border=BT; c4.number_format='#,##0'
        if alt: c4.fill=F(alt)

        # เพิ่ม/ลด vs เดือนก่อน
        if ri_off > 1:
            c6=ws.cell(row=ri,column=6)
            c6.value=f'=(B{ri}-B{ri-1})/B{ri-1}'
            c6.font=Font(name="Arial",size=9); c6.alignment=A(); c6.border=BT
            c6.number_format='+0.0%;-0.0%;0.0%'
            if alt: c6.fill=F(alt)

    # Total row
    tr = 8 + len(months)
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True)
    c=ws.cell(row=tr,column=2); c.value=f"=B5"
    c.font=Font(name="Arial",bold=True,size=10); c.alignment=A()
    c.border=BM; c.fill=F("2E75B6"); c.font=Font(name="Arial",bold=True,color="FFFFFF")
    c.number_format='#,##0.00'
    W(ws,tr,3,"100%",bg="BDD7EE",b=True,fmt='0.0%')

    ws.column_dimensions['A'].width=10; ws.column_dimensions['B'].width=18
    ws.column_dimensions['C'].width=14; ws.column_dimensions['D'].width=14
    ws.column_dimensions['E'].width=14; ws.column_dimensions['F'].width=20

# ─── SHEET 2: Ranking ร้านค้า ────────────────────────────────────────────────
def build_sheet2(wb, max_row, months, store_list):
    ws = wb.create_sheet("2_Ranking_ร้านค้า")
    ws.sheet_properties.tabColor = "2E75B6"
    DR = max_row

    W(ws,1,1,"Ranking ยอดขายรายเดือน – ชื่อร้าน  (สูตรดึงจาก DATA โดยตรง)",
      bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=4+len(months))

    hdrs=['อันดับ','ชื่อร้าน']+[MTH[m] for m in months]+['รวม (บาท)','เฉลี่ย/เดือน','%ยอดรวม']
    for ci,h in enumerate(hdrs,1):
        W(ws,3,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height=28

    for ri_off,store in enumerate(store_list,1):
        ri = ri_off + 3
        ws.row_dimensions[ri].height=15
        alt = "F0F4FF" if ri_off%2==0 else None

        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10))
        W(ws,ri,2,store,bg=alt,h="left",b=(ri_off<=5),sz=9)

        for ci,m in enumerate(months,3):
            c=ws.cell(row=ri,column=ci)
            c.value=(f'=SUMPRODUCT((DATA!{CBB}{DAT}:{CBB}{DR}=B{ri})'
                     f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                     f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")'
                     f'*DATA!{CAW}{DAT}:{CAW}{DR})')
            c.font=Font(name="Arial",size=8); c.alignment=A()
            c.border=BT; c.number_format='#,##0'
            if alt: c.fill=F(alt)

        # รวม
        first_m_col = gc(3); last_m_col = gc(2+len(months))
        ct = ws.cell(row=ri,column=3+len(months))
        ct.value=f'=SUM({first_m_col}{ri}:{last_m_col}{ri})'
        ct.font=Font(name="Arial",bold=(ri_off<=10),size=9); ct.alignment=A()
        ct.border=BT; ct.number_format='#,##0.00'; ct.fill=F("BDD7EE")

        # เฉลี่ย
        ca=ws.cell(row=ri,column=4+len(months))
        ca.value=f'=AVERAGE({first_m_col}{ri}:{last_m_col}{ri})'
        ca.font=Font(name="Arial",size=8); ca.alignment=A()
        ca.border=BT; ca.number_format='#,##0'
        if alt: ca.fill=F(alt)

        # %
        cp=ws.cell(row=ri,column=5+len(months))
        cp.value=f'={gc(3+len(months))}{ri}/\'1_ภาพรวมยอดขาย\'!B5'
        cp.font=Font(name="Arial",size=8); cp.alignment=A()
        cp.border=BT; cp.number_format='0.00%'
        if alt: cp.fill=F(alt)

    # Total row
    tr = len(store_list) + 4
    W(ws,tr,2,"รวมทั้งหมด",bg="2E75B6",fg="FFFFFF",b=True,h="left")
    for ci,m in enumerate(months,3):
        c=ws.cell(row=tr,column=ci)
        c.value=f'=SUM({gc(ci)}4:{gc(ci)}{tr-1})'
        c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.alignment=A()
        c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0'
    ct=ws.cell(row=tr,column=3+len(months))
    ct.value=f"='1_ภาพรวมยอดขาย'!B5"
    ct.font=Font(name="Arial",bold=True,color="FFFFFF"); ct.alignment=A()
    ct.border=BM; ct.fill=F("1F4E79"); ct.number_format='#,##0.00'

    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=30
    for i in range(3,3+len(months)+4): ws.column_dimensions[gc(i)].width=12
    ws.freeze_panes="C4"

# ─── SHEET 3: Ranking รหัสลูกค้า ─────────────────────────────────────────────
def build_sheet3(wb, max_row, months, code_list, codes_df):
    ws = wb.create_sheet("3_Ranking_รหัสลูกค้า")
    ws.sheet_properties.tabColor = "375623"
    DR = max_row

    W(ws,1,1,"Ranking ยอดขายรายเดือน – รหัสลูกค้า (สูตรดึงจาก DATA โดยตรง)",
      bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=5+len(months))

    hdrs=['อันดับ','ชื่อร้าน','รหัสลูกค้า']+[MTH[m] for m in months]+['รวม','เฉลี่ย','%ยอดรวม']
    for ci,h in enumerate(hdrs,1):
        W(ws,3,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height=28

    for ri_off,(code,row) in enumerate(codes_df.iterrows(),1):
        ri = ri_off+3; alt="F0F4FF" if ri_off%2==0 else None
        ws.row_dimensions[ri].height=14
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10))
        W(ws,ri,2,row['ร้าน'],bg=alt,h="left",sz=8)
        W(ws,ri,3,code,bg=alt,h="left",sz=8)

        for ci,m in enumerate(months,4):
            c=ws.cell(row=ri,column=ci)
            c.value=(f'=SUMPRODUCT((DATA!{CBC}{DAT}:{CBC}{DR}=C{ri})'
                     f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                     f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")'
                     f'*DATA!{CAW}{DAT}:{CAW}{DR})')
            c.font=Font(name="Arial",size=8); c.alignment=A(); c.border=BT; c.number_format='#,##0'
            if alt: c.fill=F(alt)

        fm=gc(4); lm=gc(3+len(months))
        ct=ws.cell(row=ri,column=4+len(months))
        ct.value=f'=SUM({fm}{ri}:{lm}{ri})'
        ct.font=Font(name="Arial",bold=(ri_off<=10),size=9); ct.alignment=A()
        ct.border=BT; ct.number_format='#,##0.00'; ct.fill=F("BDD7EE")

        ca=ws.cell(row=ri,column=5+len(months))
        ca.value=f'=AVERAGE({fm}{ri}:{lm}{ri})'
        ca.font=Font(name="Arial",size=8); ca.alignment=A(); ca.border=BT; ca.number_format='#,##0'
        if alt: ca.fill=F(alt)

        cp=ws.cell(row=ri,column=6+len(months))
        cp.value=f'={gc(4+len(months))}{ri}/\'1_ภาพรวมยอดขาย\'!B5'
        cp.font=Font(name="Arial",size=8); cp.alignment=A(); cp.border=BT; cp.number_format='0.00%'
        if alt: cp.fill=F(alt)

    tr=len(code_list)+4
    W(ws,tr,3,"รวมทั้งหมด",bg="2E75B6",fg="FFFFFF",b=True,h="left")
    for ci,m in enumerate(months,4):
        c=ws.cell(row=tr,column=ci)
        c.value=f'=SUM({gc(ci)}4:{gc(ci)}{tr-1})'
        c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.alignment=A()
        c.border=BM; c.fill=F("2E75B6"); c.number_format='#,##0'
    ct=ws.cell(row=tr,column=4+len(months))
    ct.value=f"='1_ภาพรวมยอดขาย'!B5"
    ct.font=Font(name="Arial",bold=True,color="FFFFFF"); ct.alignment=A()
    ct.border=BM; ct.fill=F("1F4E79"); ct.number_format='#,##0.00'

    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=26
    ws.column_dimensions['C'].width=32
    for i in range(4,4+len(months)+4): ws.column_dimensions[gc(i)].width=12
    ws.freeze_panes="D4"

# ─── SHEET 5: เป้าหมาย PLC & Recco ──────────────────────────────────────────
def build_sheet5(wb, max_row, months, plc_target, plc_deadline, recco_target, recco_deadline):
    ws = wb.create_sheet("5_เป้าหมาย_PLC_Recco")
    ws.sheet_properties.tabColor = "C00000"
    DR = max_row

    W(ws,1,1,"ติดตามเป้าหมายยอดขาย (สูตรดึงจาก DATA)",
      bg="C00000",fg="FFFFFF",b=True,sz=13,h="left",merge_to=12)

    plc_store  = "Pet Lover Centre"
    recco_name = "บริษัท เรคโค เพ็ท จำกัด"

    def block(sc, store_name_or_formula, target, deadline, use_bb=True):
        col_ref = CBB if use_bb else CE   # BB=ชื่อร้าน or E=ชื่อลูกค้า
        W(ws,3,sc,store_name_or_formula if isinstance(store_name_or_formula,str) and not store_name_or_formula.startswith('=') else store_name_or_formula,
          bg="2E75B6",fg="FFFFFF",b=True,sz=11,merge_to=sc+5)

        # Target cell (editable)
        W(ws,4,sc,"เป้าหมาย (บาท)",bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
        tc=ws.cell(row=4,column=sc+2); tc.value=target
        tc.font=Font(name="Arial",bold=True,size=11); tc.alignment=A()
        tc.border=BT; tc.fill=F("FFFFFF"); tc.number_format='#,##0'
        ws.merge_cells(start_row=4,start_column=sc+2,end_row=4,end_column=sc+5)
        tgt_ref = f"{gc(sc+2)}4"

        # Actual = SUMPRODUCT
        if use_bb:
            actual_f = (f'=SUMPRODUCT((DATA!{CBB}{DAT}:{CBB}{DR}="{store_name_or_formula}")'
                        f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
        else:
            # Recco uses ชื่อลูกค้า directly
            actual_f = (f'=SUMPRODUCT((DATA!{CE}{DAT}:{CE}{DR}="{store_name_or_formula}")'
                        f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')

        W(ws,5,sc,"ยอดปัจจุบัน (บาท)",bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
        ac=ws.cell(row=5,column=sc+2); ac.value=actual_f
        ac.font=Font(name="Arial",bold=True,size=11); ac.alignment=A()
        ac.border=BT; ac.fill=F("DEEAF1"); ac.number_format='#,##0.00'
        ws.merge_cells(start_row=5,start_column=sc+2,end_row=5,end_column=sc+5)
        act_ref = f"{gc(sc+2)}5"

        W(ws,6,sc,"ยอดที่เหลือ (บาท)",bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
        rc=ws.cell(row=6,column=sc+2)
        rc.value=f'=MAX(0,{tgt_ref}-{act_ref})'
        rc.font=Font(name="Arial",bold=True,size=11); rc.alignment=A()
        rc.border=BT; rc.fill=F("FCE4D6"); rc.number_format='#,##0.00'
        ws.merge_cells(start_row=6,start_column=sc+2,end_row=6,end_column=sc+5)

        W(ws,7,sc,"% บรรลุเป้า",bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
        pc=ws.cell(row=7,column=sc+2)
        pc.value=f'=IFERROR({act_ref}/{tgt_ref},0)'
        pc.font=Font(name="Arial",bold=True,size=12); pc.alignment=A()
        pc.border=BT; pc.fill=F("FFF2CC"); pc.number_format='0.0%'
        ws.merge_cells(start_row=7,start_column=sc+2,end_row=7,end_column=sc+5)

        W(ws,8,sc,"กำหนดเวลา",bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
        dc=ws.cell(row=8,column=sc+2); dc.value=deadline
        dc.font=Font(name="Arial",size=10); dc.alignment=A()
        dc.border=BT; dc.fill=F("FFFFFF")
        ws.merge_cells(start_row=8,start_column=sc+2,end_row=8,end_column=sc+5)

        # Monthly breakdown
        W(ws,10,sc,"เดือน",bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        W(ws,10,sc+1,"ยอดขาย (บาท)",bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        W(ws,10,sc+2,"% ของเป้า",bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        ws.row_dimensions[10].height=18

        for ri2,m in enumerate(months,11):
            if use_bb:
                mv_f = (f'=SUMPRODUCT((DATA!{CBB}{DAT}:{CBB}{DR}="{store_name_or_formula}")'
                        f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                        f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            else:
                mv_f = (f'=SUMPRODUCT((DATA!{CE}{DAT}:{CE}{DR}="{store_name_or_formula}")'
                        f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                        f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            alt2="F0F4FF" if ri2%2==0 else None
            W(ws,ri2,sc,MTH[m],bg=alt2,sz=8)
            mv_c=ws.cell(row=ri2,column=sc+1)
            mv_c.value=mv_f; mv_c.font=Font(name="Arial",size=8)
            mv_c.alignment=A(); mv_c.border=BT; mv_c.number_format='#,##0.00'
            if alt2: mv_c.fill=F(alt2)
            pc2=ws.cell(row=ri2,column=sc+2)
            pc2.value=f'=IFERROR({gc(sc+1)}{ri2}/{tgt_ref},0)'
            pc2.font=Font(name="Arial",size=8); pc2.alignment=A()
            pc2.border=BT; pc2.number_format='0.0%'
            if alt2: pc2.fill=F(alt2)

        sr=11+len(months)
        W(ws,sr,sc,"สะสม",bg="BDD7EE",b=True,sz=9)
        sc2=ws.cell(row=sr,column=sc+1)
        sc2.value=f'={act_ref}'
        sc2.font=Font(name="Arial",bold=True); sc2.alignment=A()
        sc2.border=BM; sc2.fill=F("BDD7EE"); sc2.number_format='#,##0.00'
        sp2=ws.cell(row=sr,column=sc+2)
        sp2.value=f'=IFERROR({act_ref}/{tgt_ref},0)'
        sp2.font=Font(name="Arial",bold=True); sp2.alignment=A()
        sp2.border=BM; sp2.fill=F("BDD7EE"); sp2.number_format='0.0%'

    block(1,  plc_store,  plc_target,  plc_deadline,  use_bb=True)
    block(8,  recco_name, recco_target, recco_deadline, use_bb=False)

    for i in range(1,15): ws.column_dimensions[gc(i)].width=14

# ─── SHEET 6: ภาพรวมสินค้า ──────────────────────────────────────────────────
def build_sheet6(wb, max_row, df):
    ws = wb.create_sheet("6_ภาพรวมสินค้า")
    ws.sheet_properties.tabColor = "7030A0"
    DR = max_row
    W(ws,1,1,"ภาพรวมยอดขายสินค้า – แยกหมวดหมู่ (สูตรดึงจาก DATA)",
      bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=5)

    # Get unique categories from Python (static list)
    df2 = df.copy()
    df2['ราคารวม'] = pd.to_numeric(df2['ราคารวม'],errors='coerce').fillna(0)
    df2 = df2[df2['สถานะรายการ']=='สำเร็จ']
    cats = df2.groupby('หมวดหมู่')['ราคารวม'].sum().sort_values(ascending=False).index.tolist()

    CAT_CLR={'Dog Food':'D6E4F7','Cat Food':'E8D5F5','Supplement':'D9F0E0',
             'Healthy Snack':'FFF2CC','Deserve Life':'FFE6D9','RAW MATERIAL':'F2F2F2','Main SKU':'F2F2F2'}

    hdrs=['หมวดหมู่','ยอดขาย (บาท)','% ของยอดรวม','จำนวนชิ้น','จำนวน SKU']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="3B1F78",fg="FFFFFF",b=True,sz=9)
    ws.row_dimensions[3].height=22

    for ri_off,cat in enumerate(cats,1):
        ri=ri_off+3; cbg=CAT_CLR.get(cat,'F2F2F2')
        W(ws,ri,1,cat,bg=cbg,h="left",b=True,sz=9)

        # ยอดขาย
        cv=ws.cell(row=ri,column=2)
        cv.value=(f'=SUMPRODUCT((DATA!{CAY}{DAT}:{CAY}{DR}=A{ri})'
                  f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
        cv.font=Font(name="Arial",size=9); cv.alignment=A()
        cv.border=BT; cv.number_format='#,##0.00'

        # %
        cp=ws.cell(row=ri,column=3)
        cp.value=f"=B{ri}/'1_ภาพรวมยอดขาย'!B5"
        cp.font=Font(name="Arial",size=9); cp.alignment=A()
        cp.border=BT; cp.number_format='0.00%'

        # จำนวนชิ้น
        cq=ws.cell(row=ri,column=4)
        cq.value=(f'=SUMPRODUCT((DATA!{CAY}{DAT}:{CAY}{DR}=A{ri})'
                  f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAT}{DAT}:{CAT}{DR})')
        cq.font=Font(name="Arial",size=9); cq.alignment=A()
        cq.border=BT; cq.number_format='#,##0'

    tr=len(cats)+4
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True,h="left")
    c=ws.cell(row=tr,column=2); c.value=f"='1_ภาพรวมยอดขาย'!B5"
    c.font=Font(name="Arial",bold=True); c.alignment=A()
    c.border=BM; c.fill=F("2E75B6"); c.font=Font(name="Arial",bold=True,color="FFFFFF")
    c.number_format='#,##0.00'
    ct=ws.cell(row=tr,column=3); ct.value="=1"; ct.number_format='0.0%'
    ct.font=Font(name="Arial",bold=True); ct.alignment=A(); ct.border=BT; ct.fill=F("BDD7EE")

    ws.column_dimensions['A'].width=18; ws.column_dimensions['B'].width=16
    ws.column_dimensions['C'].width=14; ws.column_dimensions['D'].width=13

# ─── SHEET 12: เป้าหมายรายเดือน ─────────────────────────────────────────────
def build_sheet12(wb, max_row, months, monthly_targets):
    ws = wb.create_sheet("12_เป้าหมายรายเดือน")
    ws.sheet_properties.tabColor = "FF0000"
    DR = max_row
    all_months = sorted(set(list(months)+list(monthly_targets.keys())))

    W(ws,1,1,"🎯 เป้าหมายยอดขายรายเดือน vs ยอดจริง (สูตรดึงจาก DATA)",
      bg="C00000",fg="FFFFFF",b=True,sz=13,h="left",merge_to=8)
    ws.row_dimensions[1].height=26

    # ── Target input table (editable) ──
    W(ws,3,1,"📝 ตั้งค่าเป้าหมายรายเดือน (แก้ไขได้โดยตรง)",
      bg="2E75B6",fg="FFFFFF",b=True,sz=9,h="left",merge_to=8)
    for ci,m in enumerate(all_months,1):
        W(ws,4,ci,MTH[m],bg="1F4E79",fg="FFFFFF",b=True,sz=8)
        tc=ws.cell(row=5,column=ci)
        tc.value=monthly_targets.get(m,0)
        tc.font=Font(name="Arial",bold=True,size=10); tc.alignment=A()
        tc.border=BM; tc.fill=F("FFFDE7"); tc.number_format='#,##0'
    ws.row_dimensions[5].height=20
    W(ws,5,len(all_months)+1,"← แก้ตรงนี้ได้เลย",
      bg=None,fg="C00000",b=True,sz=8,h="left",border=Border())

    # ── Results table ──
    hdrs=["เดือน","เป้าหมาย","ยอดจริง","% Achievement","Gap","สถานะ","จำนวนร้าน"]
    for ci,h in enumerate(hdrs,1):
        W(ws,7,ci,h,bg="C00000",fg="FFFFFF",b=True,sz=9,wrap=True)
    ws.row_dimensions[7].height=28

    for ri_off,m in enumerate(all_months,1):
        ri=ri_off+7; alt="F9F9F9" if ri_off%2==0 else None
        m_col = gc(ri_off)   # col in target row (row 5)
        W(ws,ri,1,MTH[m],bg=alt,b=True,sz=10)

        # เป้า → ดึงจาก row 5
        W(ws,ri,2,f'={m_col}5',bg=alt,fmt='#,##0',sz=9)

        # ยอดจริง → SUMPRODUCT
        if m in months:
            c3=ws.cell(row=ri,column=3)
            c3.value=(f'=SUMPRODUCT((DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                      f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ")*DATA!{CAW}{DAT}:{CAW}{DR})')
            c3.font=Font(name="Arial",size=9); c3.alignment=A()
            c3.border=BT; c3.number_format='#,##0.00'
            if alt: c3.fill=F(alt)
        else:
            W(ws,ri,3,"(ยังไม่มีข้อมูล)",bg="F5F5F5",fg="AAAAAA",sz=8,italic=True)

        # % Achievement
        if m in months:
            c4=ws.cell(row=ri,column=4)
            c4.value=f'=IFERROR(C{ri}/B{ri},0)'
            c4.font=Font(name="Arial",bold=True,size=9); c4.alignment=A()
            c4.border=BT; c4.number_format='0.0%'

            # Gap
            c5=ws.cell(row=ri,column=5)
            c5.value=f'=C{ri}-B{ri}'
            c5.font=Font(name="Arial",size=9); c5.alignment=A()
            c5.border=BT; c5.number_format='+#,##0;-#,##0;0'

            # สถานะ (nested IF)
            c6=ws.cell(row=ri,column=6)
            c6.value=f'=IF(D{ri}>=1,"✅ ถึงเป้า",IF(D{ri}>=0.9,"🟡 ใกล้เป้า",IF(D{ri}>=0.7,"🟠 ต่ำกว่าเป้า","🔴 ต่ำมาก")))'
            c6.font=Font(name="Arial",bold=True,size=9); c6.alignment=A(); c6.border=BT

            # จำนวนร้าน
            c7=ws.cell(row=ri,column=7)
            c7.value=(f'=SUMPRODUCT(1/COUNTIFS(DATA!{CBB}{DAT}:{CBB}{DR},DATA!{CBB}{DAT}:{CBB}{DR},'
                      f'DATA!{CBA}{DAT}:{CBA}{DR},{m},'
                      f'DATA!{CAK}{DAT}:{CAK}{DR},"สำเร็จ")'
                      f'*(DATA!{CBA}{DAT}:{CBA}{DR}={m})'
                      f'*(DATA!{CAK}{DAT}:{CAK}{DR}="สำเร็จ"))')
            c7.font=Font(name="Arial",size=9); c7.alignment=A(); c7.border=BT

    # Grand total row
    tr=len(all_months)+8
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True)
    W(ws,tr,2,f'=SUM(B8:B{tr-1})',bg="BDD7EE",b=True,fmt='#,##0')
    W(ws,tr,3,f'=SUM(C8:C{tr-1})',bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0.00')
    W(ws,tr,4,f'=IFERROR(C{tr}/B{tr},0)',bg="BDD7EE",b=True,fmt='0.0%')
    W(ws,tr,5,f'=C{tr}-B{tr}',bg="BDD7EE",b=True,fmt='+#,##0;-#,##0;0')

    for i in range(1,9): ws.column_dimensions[gc(i)].width=15
    ws.freeze_panes="A8"

# ─── MAIN ENTRY POINT ────────────────────────────────────────────────────────
def build_dashboard(file_bytes, plc_target=1_250_000, plc_deadline="30 มิถุนายน 2569",
                    recco_target=1_000_000, recco_deadline="31 ธันวาคม 2569",
                    monthly_targets=None, progress_cb=None):
    def upd(pct, msg):
        if progress_cb: progress_cb(pct, msg)

    upd(8, "📥 โหลดข้อมูล...")
    src_wb, df, c1, c2 = load_raw(file_bytes)
    n2s      = build_name_map(c1)
    code_map = build_code_map(c2)

    df2 = df.copy()
    for col in ['ชื่อลูกค้า','รหัสลูกค้า','สถานะรายการ']:
        if col in df2.columns: df2[col] = df2[col].astype(str).str.strip()
    df2['ราคารวม'] = pd.to_numeric(df2.get('ราคารวม',pd.Series(dtype=float)), errors='coerce').fillna(0)
    df2['date']    = pd.to_datetime(df2.get('วันที่ทำรายการ'), dayfirst=True, errors='coerce')
    df2['month']   = df2['date'].dt.month
    months = sorted(df2['month'].dropna().unique().astype(int))

    wb = Workbook()
    wb.remove(wb.active)

    upd(15, "📋 สร้าง lookup tables...")
    build_cond1_sheet(wb, n2s)
    build_cond2_sheet(wb, code_map)

    upd(22, "📥 Copy ข้อมูลดิบ + helper formulas...")
    max_row = build_data_sheet(wb, src_wb, n2s, code_map)

    upd(32, "📊 สร้างรายการ unique stores/codes...")
    store_list, code_list, codes_df = build_store_list_sheet(wb, df2, n2s, code_map)

    upd(40, "📊 ภาพรวมยอดขาย...")
    build_sheet1(wb, max_row, months)

    upd(50, "🏪 Ranking ร้านค้า...")
    build_sheet2(wb, max_row, months, store_list)

    upd(58, "🔑 Ranking รหัสลูกค้า...")
    build_sheet3(wb, max_row, months, code_list, codes_df)

    upd(65, "🎯 เป้าหมาย PLC & Recco...")
    build_sheet5(wb, max_row, months, plc_target, plc_deadline, recco_target, recco_deadline)

    upd(72, "📦 ภาพรวมสินค้า...")
    build_sheet6(wb, max_row, df2)

    if monthly_targets:
        upd(82, "🎯 เป้าหมายรายเดือน...")
        build_sheet12(wb, max_row, months, monthly_targets)

    upd(95, "💾 บันทึกไฟล์...")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
